"""
Excel 파일 생성 모듈
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side

# =============================================================================
# 내부 헬퍼 함수
# =============================================================================

def create_scenario_dataframe(scenario_data, total_budget):
    """
    시나리오 데이터를 DataFrame으로 변환
    
    Args:
        scenario_data: 시나리오 매체 리스트
        total_budget: 총 예산
    
    Returns:
        pandas DataFrame
    """
    data = []
    
    for media in scenario_data:
        try:
            row = {
                '매체명': media.get('name', 'N/A'),
                '카테고리': media.get('category', 'N/A'),
                '예산': int(media.get('media_budget', 0)),
                '예산비중': media.get('budget_ratio', 0),
                'CPM': int(media.get('cpm', 0)),
                '예상노출': int(media.get('estimated_impressions', 0)),
                '예상클릭': int(media.get('estimated_clicks', 0)),
                'CTR': media.get('ctr', 0),
                'CPC': media.get('cpc', 0),
                '예상전환': round(media.get('estimated_conversions_adjusted', 0), 1),
                'CVR': media.get('cvr', 0),
                'CPA': int(media.get('cpa_adjusted', 0)),
                'ROAS': round(media.get('roas_adjusted', 0), 1)
            }
            data.append(row)
        except Exception as e:
            # 매체별 계산 실패 시 안전하게 처리
            row = {
                '매체명': media.get('name', 'N/A'),
                '카테고리': media.get('category', 'N/A'),
                '예산': 0,
                '예산비중': 0,
                'CPM': 0,
                '예상노출': 0,
                '예상클릭': 0,
                'CTR': 0,
                'CPC': 0,
                '예상전환': 0,
                'CVR': 0,
                'CPA': 0,
                'ROAS': 0
            }
            data.append(row)
    
    return pd.DataFrame(data)

# =============================================================================
# Excel 파일 생성
# =============================================================================

def create_excel_download(scenarios, budget, mode_name="미디어믹스", summary_df=None, extra_data=None):
    """
    공통 Excel 다운로드 파일 생성 (최적화 버전)
    
    최적화 내용:
    - 테두리: 헤더 + 합계행만 적용 (데이터 행 생략)
    - 퍼센트 변환: DataFrame 단계에서 일괄 처리
    - 열 너비: 한 번에 설정
    - 셀 루프 최소화
    
    Args:
        scenarios: 시나리오 데이터 딕셔너리 {'conservative': [], 'base': [], 'aggressive': []}
        budget: 총 예산
        mode_name: 파일명에 사용할 모드 이름 (기본: "미디어믹스")
        summary_df: 추가 요약 데이터프레임 (옵션)
        extra_data: 추가 시트 데이터 딕셔너리 {sheet_name: dataframe} (옵션)
    
    Returns:
        BytesIO: Excel 파일 데이터
        str: 파일명
    """
    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 시나리오별 시트
        for scenario_key, sheet_name in [('conservative', '보수안(-5%)'), 
                                        ('base', '기본안'), 
                                        ('aggressive', '공격안(+10%)')]:
            if scenario_key in scenarios:
                df = create_scenario_dataframe(scenarios[scenario_key], budget)
                
                # 퍼센티지 컬럼을 DataFrame 단계에서 변환
                percentage_columns = ['예산비중', 'CTR', 'CVR', 'ROAS']
                for col in percentage_columns:
                    if col in df.columns:
                        df[col] = df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) and x != '' else x)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 워크시트 포맷 적용
                worksheet = writer.sheets[sheet_name]
                num_rows = len(df) + 1  # 헤더 포함
                num_cols = len(df.columns)
                
                # 모든 열 너비 15로 설정 (한 번에)
                for col_idx in range(1, num_cols + 1):
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = 15
                
                # 가운데 정렬 (모든 셀)
                center_alignment = Alignment(horizontal='center', vertical='center')
                for row in range(1, num_rows + 1):
                    for col in range(1, num_cols + 1):
                        worksheet.cell(row=row, column=col).alignment = center_alignment
                
                # 테두리 적용: 헤더(row=1) + 합계행(마지막 row)만
                for col in range(1, num_cols + 1):
                    # 헤더 행
                    worksheet.cell(row=1, column=col).border = thin_border
                    # 합계 행 (마지막 행)
                    worksheet.cell(row=num_rows, column=col).border = thin_border
        
        # 시나리오 요약 시트
        if summary_df is not None:
            summary_df.to_excel(writer, sheet_name='시나리오_요약', index=False)
            
            # 요약 시트 포맷 적용
            worksheet = writer.sheets['시나리오_요약']
            num_rows = len(summary_df) + 1
            num_cols = len(summary_df.columns)
            
            # 모든 열 너비 15로 설정
            for col_idx in range(1, num_cols + 1):
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[col_letter].width = 15
            
            # 가운데 정렬
            center_alignment = Alignment(horizontal='center', vertical='center')
            for row in range(1, num_rows + 1):
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=row, column=col).alignment = center_alignment
            
            # 테두리: 헤더 + 마지막 행만
            for col in range(1, num_cols + 1):
                worksheet.cell(row=1, column=col).border = thin_border
                worksheet.cell(row=num_rows, column=col).border = thin_border
        
        # 추가 시트
        if extra_data:
            for sheet_name, df in extra_data.items():
                # 퍼센티지 컬럼 변환
                percentage_columns = ['예산비중', 'CTR', 'CVR', 'ROAS']
                for col in percentage_columns:
                    if col in df.columns:
                        df[col] = df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) and x != '' else x)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 포맷 적용
                worksheet = writer.sheets[sheet_name]
                num_rows = len(df) + 1
                num_cols = len(df.columns)
                
                # 열 너비 설정
                for col_idx in range(1, num_cols + 1):
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = 15
                
                # 가운데 정렬
                center_alignment = Alignment(horizontal='center', vertical='center')
                for row in range(1, num_rows + 1):
                    for col in range(1, num_cols + 1):
                        worksheet.cell(row=row, column=col).alignment = center_alignment
                
                # 테두리: 헤더 + 마지막 행만
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=1, column=col).border = thin_border
                    worksheet.cell(row=num_rows, column=col).border = thin_border
    
    output.seek(0)
    
    now = datetime.now()
    filename = f"{mode_name}_{now.strftime('%Y년%m월_%Y%m%d_%H%M%S')}.xlsx"
    
    return output, filename

