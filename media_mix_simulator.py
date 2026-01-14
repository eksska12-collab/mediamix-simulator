# -*- coding: utf-8 -*-
"""
미디어믹스 시뮬레이터
Media Mix Simulator
"""

import pandas as pd
import json
import os
import sys
from datetime import datetime
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import Workbook

# =============================================================================
# JSON 데이터 로드
# =============================================================================

def load_benchmarks_json():
    """
    벤치마크 데이터를 JSON 파일에서 로드
    
    Returns:
        dict: 벤치마크 데이터 전체
    """
    try:
        with open('benchmarks.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # JSON의 숫자 키를 정수로 변환 (SEASONALITY_COMMON)
        if 'SEASONALITY_COMMON' in data:
            data['SEASONALITY_COMMON'] = {
                int(k) if k.isdigit() else k: v 
                for k, v in data['SEASONALITY_COMMON'].items()
                if k != 'description'
            }
        
        # SEASONALITY도 정수 키로 변환
        if 'SEASONALITY' in data:
            data['SEASONALITY'] = {
                int(k) if k.isdigit() else k: v 
                for k, v in data['SEASONALITY'].items()
                if k != 'note'
            }
        
        return data
    except FileNotFoundError:
        print("⚠️ benchmarks.json 파일을 찾을 수 없습니다. 기본값을 사용합니다.")
        return {}
    except json.JSONDecodeError as e:
        print(f"⚠️ benchmarks.json 파일 파싱 오류: {e}")
        return {}

def load_media_categories_json():
    """
    매체 카테고리 데이터를 JSON 파일에서 로드
    
    Returns:
        dict: 매체 카테고리 데이터 전체
    """
    try:
        with open('media_categories.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print("⚠️ media_categories.json 파일을 찾을 수 없습니다. 기본값을 사용합니다.")
        return {}
    except json.JSONDecodeError as e:
        print(f"⚠️ media_categories.json 파일 파싱 오류: {e}")
        return {}

# JSON 데이터 로드
_benchmarks_data = load_benchmarks_json()
_media_data = load_media_categories_json()

# 매체 카테고리 정의 (JSON에서 로드)
MEDIA_CATEGORIES = _media_data.get('MEDIA_CATEGORIES_CLI', {})

# =============================================================================
# 벤치마크 데이터 (JSON에서 로드)
# =============================================================================

# 업종별 Base Metrics (JSON에서 로드)
INDUSTRY_BASE_METRICS = _benchmarks_data.get('INDUSTRY_BASE_METRICS', {})

# 계절성 공통 보정 계수 (JSON에서 로드)
SEASONALITY_COMMON = _benchmarks_data.get('SEASONALITY_COMMON', {})

# 업종별 계절성 가중치 (JSON에서 로드)
INDUSTRY_SEASON_WEIGHT = _benchmarks_data.get('INDUSTRY_SEASON_WEIGHT', {})

# 매체별 성과 배율 (JSON에서 로드)
MEDIA_MULTIPLIERS = _benchmarks_data.get('MEDIA_MULTIPLIERS', {})

# 업종별 매체 효율 벤치마크 DB (JSON에서 로드)
BENCHMARKS = _benchmarks_data.get('BENCHMARKS', {})

# 계절성 보정 테이블 (JSON에서 로드, 기존 호환용)
SEASONALITY = _benchmarks_data.get('SEASONALITY', {})


# =============================================================================
# 보정 계산 함수들
# =============================================================================

def calculate_budget_competition_factor(budget):
    """
    예산 규모별 경쟁도 보정 계수 계산 (CPC 전용)
    
    예산이 클수록 경쟁이 심한 키워드를 입찰하게 되어 CPC가 상승하는 효과 반영
    
    Args:
        budget: 총 예산 (원)
    
    Returns:
        float: CPC 보정 계수
            - 1천만원 미만: 0.90 (중소 예산, 경쟁 낮은 롱테일 키워드)
            - 1천만원~5천만원: 1.00 (기준)
            - 5천만원~1억: 1.10 (경쟁 키워드 진입)
            - 1억 이상: 1.20 (프리미엄 키워드 경쟁)
    """
    if budget < 10_000_000:      # 1천만원 미만
        return 0.90
    elif budget < 50_000_000:    # 5천만원 미만
        return 1.00
    elif budget < 100_000_000:   # 1억 미만
        return 1.10
    else:                        # 1억 이상
        return 1.20


def apply_adjustments(industry, month, budget, base_metrics):
    """
    업종/계절성/예산 보정을 통합 적용
    
    보정 적용 순서:
    1. 계절성 공통 보정 (SEASONALITY_COMMON)
    2. 업종별 계절성 가중치 (INDUSTRY_SEASON_WEIGHT)
    3. 예산 규모 경쟁도 보정 (CPC만)
    
    Args:
        industry (str): 업종명 ('보험', '금융', '패션', 'IT/테크')
        month (int): 월 (1-12)
        budget (float): 총 예산 (원)
        base_metrics (dict): {'CTR': x, 'CPC': y, 'CVR': z}
    
    Returns:
        dict: 보정된 지표
            {
                'CTR': 보정된 클릭률,
                'CVR': 보정된 전환율,
                'CPC': 보정된 클릭당비용
            }
    
    Example:
        >>> base = {'CTR': 0.01, 'CPC': 1000, 'CVR': 0.02}
        >>> adjusted = apply_adjustments('보험', 12, 50000000, base)
        >>> # 12월 성수기 + 보험 업종 특성 반영된 지표 반환
    """
    # 1. 계절성 공통 보정
    season_factor = SEASONALITY_COMMON.get(month, 1.0)
    
    # 2. 업종별 계절성 가중치 적용
    if industry in INDUSTRY_SEASON_WEIGHT:
        industry_season = INDUSTRY_SEASON_WEIGHT[industry]
        if month in industry_season['high_months']:
            # 성수기: 공통 보정에 추가 배율 적용
            season_factor *= industry_season['high_multiplier']
        elif month in industry_season['low_months']:
            # 비수기: 공통 보정에 감소 배율 적용
            season_factor *= industry_season['low_multiplier']
    
    # 3. 예산 규모 경쟁도 보정 (CPC만)
    competition_factor = calculate_budget_competition_factor(budget)
    
    # 4. 보정 적용
    adjusted = {
        'CTR': base_metrics['CTR'] * season_factor,       # 계절성 영향
        'CVR': base_metrics['CVR'] * season_factor,       # 계절성 영향
        'CPC': base_metrics['CPC'] * competition_factor   # 예산 규모 영향
    }
    
    return adjusted


def calculate_performance(budget, adjusted_metrics):
    """
    보정된 지표로 예상 성과 계산
    
    계산 공식:
    - 클릭수 = 예산 / CPC
    - 노출수 = 클릭수 / CTR
    - 전환수 = 클릭수 × CVR
    - CPA = 예산 / 전환수
    
    Args:
        budget (float): 예산 (원)
        adjusted_metrics (dict): 보정된 지표
            {'CTR': x, 'CPC': y, 'CVR': z}
    
    Returns:
        dict: 예상 성과
            {
                'impressions': 노출수,
                'clicks': 클릭수,
                'conversions': 전환수,
                'cpa': 전환당비용,
                'ctr': 클릭률(%),
                'cvr': 전환율(%),
                'cpc': 클릭당비용
            }
    
    Raises:
        ValueError: 잘못된 지표 값 (0 이하)
        ZeroDivisionError: 계산 불가능한 값
    """
    try:
        cpc = adjusted_metrics['CPC']
        ctr = adjusted_metrics['CTR']
        cvr = adjusted_metrics['CVR']
        
        # 입력값 검증
        if cpc <= 0:
            raise ValueError(f"CPC must be greater than 0, got {cpc}")
        if ctr <= 0:
            raise ValueError(f"CTR must be greater than 0, got {ctr}")
        if cvr < 0:
            raise ValueError(f"CVR must be non-negative, got {cvr}")
        if budget <= 0:
            raise ValueError(f"Budget must be greater than 0, got {budget}")
        
        # 성과 지표 계산 (ZeroDivisionError 방지)
        clicks = budget / cpc
        impressions = clicks / ctr
        conversions = clicks * cvr
        cpa = budget / conversions if conversions > 0 else 0
        
        return {
            'impressions': round(impressions, 0),
            'clicks': round(clicks, 0),
            'conversions': round(conversions, 2),
            'cpa': round(cpa, 0),
            'ctr': round(ctr * 100, 2),   # % 변환
            'cvr': round(cvr * 100, 2),   # % 변환
            'cpc': round(cpc, 0)
        }
    
    except ZeroDivisionError as e:
        raise ZeroDivisionError(f"계산 중 0으로 나누기 오류: {e}")
    except KeyError as e:
        raise KeyError(f"필수 지표가 누락되었습니다: {e}")
    except Exception as e:
        raise Exception(f"성과 계산 중 오류 발생: {e}")


def get_media_adjusted_metrics(industry, media_key, month, budget):
    """
    특정 매체의 보정된 성과 지표 계산
    
    프로세스:
    1. 업종 Base Metrics 로드
    2. 매체별 Multiplier 적용
    3. 계절성/예산 보정 적용
    
    Args:
        industry (str): 업종명
        media_key (str): 매체 키 (예: '네이버_SA', '메타_DA')
        month (int): 월 (1-12)
        budget (float): 해당 매체 예산
    
    Returns:
        dict: 보정된 지표 또는 None (데이터 없을 시)
    
    Example:
        >>> metrics = get_media_adjusted_metrics('패션', '메타_DA', 11, 30000000)
        >>> # 패션 업종, 메타 디스플레이, 11월, 3천만원 예산의 보정 지표
    """
    # 1. 업종 Base Metrics 확인
    if industry not in INDUSTRY_BASE_METRICS:
        return None
    
    base = INDUSTRY_BASE_METRICS[industry]
    
    # 2. 매체 Multiplier 적용
    if media_key in MEDIA_MULTIPLIERS:
        multiplier = MEDIA_MULTIPLIERS[media_key]
        media_base = {
            'CTR': base['CTR'] * multiplier['CTR'],
            'CPC': base['CPC'] * multiplier['CPC'],
            'CVR': base['CVR'] * multiplier['CVR']
        }
    else:
        # Multiplier 없으면 Base 그대로 사용
        media_base = base.copy()
    
    # 3. 계절성/예산 보정 적용
    adjusted = apply_adjustments(industry, month, budget, media_base)
    
    return adjusted


def get_number_input(prompt, allow_zero=False, allow_negative=False, min_val=None, max_val=None):
    """
    숫자 입력 검증 함수
    문자 입력 시 재입력 요청
    
    Args:
        prompt: 입력 안내 메시지
        allow_zero: 0 허용 여부
        allow_negative: 음수 허용 여부
        min_val: 최소값 (None이면 제한 없음)
        max_val: 최대값 (None이면 제한 없음)
    
    Returns:
        유효한 숫자 입력값
    """
    while True:
        try:
            value = input(prompt)
            # 쉼표 제거
            value = value.replace(',', '')
            number = float(value)
            
            if not allow_negative and number < 0:
                print("❌ 음수는 입력할 수 없습니다. 다시 입력해주세요.")
                continue
            
            if not allow_zero and number == 0:
                print("❌ 0은 입력할 수 없습니다. 다시 입력해주세요.")
                continue
            
            if min_val is not None and number < min_val:
                if max_val is not None:
                    print(f"❌ {min_val}~{max_val} 사이 값을 입력하세요.")
                else:
                    print(f"❌ {min_val} 이상의 값을 입력하세요.")
                continue
            
            if max_val is not None and number > max_val:
                if min_val is not None:
                    print(f"❌ {min_val}~{max_val} 사이 값을 입력하세요.")
                else:
                    print(f"❌ {max_val} 이하의 값을 입력하세요.")
                continue
            
            return number
        except ValueError:
            print("❌ 숫자만 입력 가능합니다.")


def format_number(number):
    """숫자를 천 단위 쉼표 형식으로 변환"""
    return f"{int(number):,}"


def ensure_saved_data_folder():
    """saved_data 폴더가 없으면 생성"""
    if not os.path.exists('saved_data'):
        os.makedirs('saved_data')


def create_readme():
    """README.txt 파일 자동 생성 (없을 때만)"""
    if os.path.exists('README.txt'):
        return
    
    readme_content = """
================================================================================
미디어믹스 시뮬레이터 (Media Mix Simulator)
================================================================================

[프로그램 설명]
광고 매체별 예산 배분과 성과를 시뮬레이션하는 도구입니다.
3가지 시나리오(보수안, 기본안, 공격안)로 성과를 예측하고 Excel 파일로 결과를 저장합니다.

[실행 방법]
1. 필요한 라이브러리 설치:
   pip install -r requirements.txt

2. 프로그램 실행:
   python media_mix_simulator.py

3. 모드 선택:
   - 빠른 모드: 매체 3개 이하로 빠르게 시뮬레이션
   - 상세 모드: 제한 없이 여러 매체 분석
   - 불러오기: 저장된 데이터로 이어서 작업

[사용 흐름]
1. 총 예산 입력
2. 매체 선택 (카테고리별)
3. 각 매체별 상세 데이터 입력:
   - 예산 비중 (%)
   - CPC (클릭당 비용)
   - CTR (클릭률)
   - CVR (전환율)
   - 전환당 매출액
   - 예측 오차 범위
4. 시나리오 조정 폭 설정
5. 계산 및 Excel 저장

[주의사항]
- 예산 비중 합계는 반드시 100%가 되어야 합니다.
- 중간 저장 기능을 활용하여 데이터를 보관할 수 있습니다.
- 저장된 파일은 saved_data 폴더에 저장됩니다.
- 결과 Excel 파일은 현재 디렉토리에 저장됩니다.

[출력 파일]
- Excel 파일: 미디어믹스_YYYY년M월_YYYYMMDD_HHMMSS.xlsx
  * 보수안, 기본안, 공격안 시트
  * 시나리오 요약 시트
- JSON 중간 저장 파일: saved_data/[파일명].json

[매체 카테고리]
- 검색광고: 네이버, 카카오, 구글, 기타
- 디스플레이광고: 메타, 구글GDA, 카카오DA, 틱톡, 기타
- 동영상광고: 유튜브, 틱톡, 네이버TV, 기타

[문의]
프로그램 사용 중 문제가 발생하면 개발팀에 문의하세요.

================================================================================
"""
    
    with open('README.txt', 'w', encoding='utf-8') as f:
        f.write(readme_content)


def print_progress(step, total, description):
    """진행 상황 표시"""
    print(f"\n[{step}/{total}] {description}")
    print("─" * 50)


def confirm_input_data(budget, selected_media):
    """
    입력값 확인 화면
    
    Args:
        budget: 총 예산
        selected_media: 매체 리스트
    
    Returns:
        'continue': 계속 진행
        'modify': 수정
        'save': 저장 후 계속
    """
    print("\n" + "="*50)
    print("📋 입력 내용 확인")
    print("="*50)
    
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    
    print(f"\n{'매체명':<15} {'카테고리':<12} {'비중':<8} {'CPC':<10} {'CTR':<8} {'CVR':<8} {'전환매출':<12} {'오차범위'}")
    print("-"*100)
    
    for media in selected_media:
        print(f"{media['name']:<15} "
              f"{media['category']:<12} "
              f"{media['budget_ratio']:>6.2f}% "
              f"{format_number(media['cpc']):>10}원 "
              f"{media['ctr']:>6.2f}% "
              f"{media['cvr']:>6.2f}% "
              f"{format_number(media['revenue_per_cv']):>12}원 "
              f"{media['adjustment']:>+6.1f}%")
    
    print("\n" + "="*50)
    while True:
        choice = input("입력 내용이 맞습니까? (y: 계속 / n: 수정 / s: 저장): ").strip().lower()
        if choice in ['y', 'n', 's']:
            if choice == 'y':
                return 'continue'
            elif choice == 'n':
                return 'modify'
            else:
                return 'save'
        print("❌ 'y', 'n', 또는 's'를 입력해주세요.")


def save_intermediate_data(budget, selected_media):
    """
    중간 저장 기능
    
    Args:
        budget: 총 예산
        selected_media: 매체 리스트
    
    Returns:
        저장 성공 여부 (bool)
    """
    print("\n" + "="*50)
    save_choice = input("중간 저장하시겠습니까? (y/n): ").strip().lower()
    
    if save_choice != 'y':
        return False
    
    # saved_data 폴더 확인/생성
    ensure_saved_data_folder()
    
    # 기본 파일명 생성
    now = datetime.now()
    default_filename = f"temp_{now.strftime('%Y%m%d_%H%M%S')}"
    
    filename_input = input(f"파일명 입력 (엔터={default_filename}): ").strip()
    filename = filename_input if filename_input else default_filename
    
    # .json 확장자 추가 (없으면)
    if not filename.endswith('.json'):
        filename += '.json'
    
    # 데이터 구조화
    data = {
        'total_budget': budget,
        'media_list': selected_media,
        'saved_at': now.strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # JSON 파일로 저장
    filepath = os.path.join('saved_data', filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 저장 완료: {filepath}")
    return True


def load_saved_data_file():
    """
    저장된 데이터 불러오기
    
    Returns:
        (budget, selected_media) 튜플 또는 None
    """
    # saved_data 폴더 확인
    if not os.path.exists('saved_data'):
        print("\n⚠️ 저장된 데이터가 없습니다.")
        return None
    
    # JSON 파일 목록 가져오기
    json_files = [f for f in os.listdir('saved_data') if f.endswith('.json')]
    
    if not json_files:
        print("\n⚠️ 저장된 데이터가 없습니다.")
        return None
    
    # 파일 목록 출력
    print("\n" + "="*50)
    print("📁 저장된 데이터 목록")
    print("="*50)
    
    for i, filename in enumerate(json_files, 1):
        filepath = os.path.join('saved_data', filename)
        # 파일 수정 시간 가져오기
        mtime = os.path.getmtime(filepath)
        mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
        print(f"  {i}. {filename} (저장일시: {mtime_str})")
    
    # 파일 선택
    while True:
        try:
            choice = int(get_number_input(f"\n파일 선택 (1-{len(json_files)}, 0=취소): ", allow_zero=True, allow_negative=False))
            
            if choice == 0:
                return None
            
            if 1 <= choice <= len(json_files):
                selected_file = json_files[choice - 1]
                break
            else:
                print(f"❌ 1-{len(json_files)} 사이의 숫자를 입력해주세요.")
        except:
            print("❌ 올바른 숫자를 입력해주세요.")
    
    # 파일 로드
    filepath = os.path.join('saved_data', selected_file)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        budget = data.get('total_budget')
        selected_media = data.get('media_list')
        saved_at = data.get('saved_at', '알 수 없음')
        
        # 불러온 데이터 요약 출력
        print("\n" + "="*50)
        print("✅ 데이터 로드 완료")
        print("="*50)
        print(f"\n💾 저장일시: {saved_at}")
        print(f"💰 총 예산: {format_number(budget)}원")
        print(f"📱 매체 수: {len(selected_media)}개")
        print("\n📋 매체 목록:")
        for i, media in enumerate(selected_media, 1):
            budget_info = f", 비중: {media['budget_ratio']:.2f}%" if 'budget_ratio' in media else ""
            print(f"  {i}. {media['name']} ({media['category']}{budget_info})")
        
        # 이어서 진행 여부 확인
        print("\n" + "-"*50)
        while True:
            continue_choice = input("이어서 진행하시겠습니까? (y/n): ").strip().lower()
            if continue_choice in ['y', 'n']:
                break
            print("❌ 'y' 또는 'n'을 입력해주세요.")
        
        if continue_choice == 'y':
            return (budget, selected_media)
        else:
            return None
    
    except Exception as e:
        print(f"❌ 파일 로드 중 오류 발생: {e}")
        return None


def modify_budget_ratio(selected_media):
    """
    예산 비중 수정 기능
    
    Args:
        selected_media: 매체 리스트
    
    Returns:
        수정된 매체 리스트
    """
    while True:
        print("\n" + "="*50)
        print("💡 예산 비중 수정")
        print("="*50)
        print("\n현재 매체 목록:")
        for i, media in enumerate(selected_media, 1):
            print(f"  {i}. {media['name']}: {media['budget_ratio']:.2f}%")
        print(f"  0. 전체 재입력")
        
        try:
            choice = int(get_number_input(f"\n수정할 매체 선택 (0-{len(selected_media)}): ", allow_zero=True, allow_negative=False))
            
            if choice == 0:
                # 전체 재입력
                return None
            
            if 1 <= choice <= len(selected_media):
                media_index = choice - 1
                media = selected_media[media_index]
                
                print(f"\n📝 {media['name']} 예산 비중 수정")
                print(f"현재 값: {media['budget_ratio']:.2f}%")
                
                while True:
                    new_ratio = get_number_input("새로운 비중 (%): ", allow_zero=True, allow_negative=False)
                    if 0 <= new_ratio <= 100:
                        old_ratio = media['budget_ratio']
                        media['budget_ratio'] = new_ratio
                        print(f"✅ {media['name']} 비중: {old_ratio:.2f}% → {new_ratio:.2f}%")
                        break
                    else:
                        print("❌ 0~100 사이의 값을 입력해주세요.")
                
                # 현재 합계 출력
                total_ratio = sum(m['budget_ratio'] for m in selected_media)
                print(f"\n📊 현재 예산 비중 합계: {total_ratio:.2f}%")
                
                # 계속 수정할지 확인
                if abs(total_ratio - 100) < 0.01:
                    print("✅ 예산 비중 합계가 100%입니다!")
                    return selected_media
                else:
                    print(f"⚠️ 합계가 {total_ratio:.2f}%입니다. 100%가 되어야 합니다.")
                    continue_modify = input("\n다른 매체도 수정하시겠습니까? (y/n): ").strip().lower()
                    if continue_modify != 'y':
                        return selected_media
            else:
                print(f"❌ 0-{len(selected_media)} 사이의 숫자를 입력해주세요.")
        
        except:
            print("❌ 올바른 숫자를 입력해주세요.")


def get_budget_input():
    """
    총 예산 입력 함수
    
    Returns:
        입력된 예산 (float)
    """
    print("\n💰 총 예산을 입력하세요")
    budget = get_number_input("예산 (원): ", allow_zero=False, allow_negative=False)
    
    # 확인 메시지 출력
    print(f"✅ 입력된 예산: {format_number(budget)}원")
    return budget


def select_media(max_count=None):
    """
    매체 선택 함수
    
    Args:
        max_count: 최대 선택 가능한 매체 수 (None이면 제한 없음)
    
    Returns:
        선택된 매체 리스트 [{'name': '네이버', 'category': '검색광고'}, ...]
    """
    selected_media = []
    category_names = list(MEDIA_CATEGORIES.keys())
    
    print("\n📱 매체 선택")
    if max_count:
        print(f"(최대 {max_count}개까지 선택 가능)")
    
    while True:
        # 최대 개수 체크
        if max_count and len(selected_media) >= max_count:
            print(f"\n⚠️ 최대 {max_count}개까지만 선택 가능합니다.")
            break
        
        # 카테고리 선택
        print("\n" + "-"*50)
        print("카테고리를 선택하세요:")
        for i, category in enumerate(category_names, 1):
            print(f"  {i}. {category}")
        
        try:
            category_choice = int(get_number_input("카테고리 선택 (1-3): ", allow_zero=False, allow_negative=False))
            if category_choice < 1 or category_choice > 3:
                print("❌ 1-3 사이의 숫자를 입력해주세요.")
                continue
        except:
            print("❌ 올바른 숫자를 입력해주세요.")
            continue
        
        selected_category = category_names[category_choice - 1]
        media_list = MEDIA_CATEGORIES[selected_category]
        
        # 매체 선택
        print(f"\n{selected_category} 매체를 선택하세요:")
        for i, media in enumerate(media_list, 1):
            print(f"  {i}. {media}")
        
        try:
            media_choice = int(get_number_input(f"매체 선택 (1-{len(media_list)}): ", allow_zero=False, allow_negative=False))
            if media_choice < 1 or media_choice > len(media_list):
                print(f"❌ 1-{len(media_list)} 사이의 숫자를 입력해주세요.")
                continue
        except:
            print("❌ 올바른 숫자를 입력해주세요.")
            continue
        
        selected_media_name = media_list[media_choice - 1]
        
        # "기타" 선택 시 직접 입력
        if selected_media_name == "기타":
            custom_name = input("매체명을 입력하세요: ").strip()
            if custom_name:
                selected_media_name = custom_name
            else:
                print("❌ 매체명을 입력해주세요.")
                continue
        
        # 중복 체크
        if any(media['name'] == selected_media_name for media in selected_media):
            print(f"⚠️ '{selected_media_name}'는 이미 선택된 매체입니다.")
            continue
        
        # 매체 추가
        selected_media.append({
            'name': selected_media_name,
            'category': selected_category
        })
        
        # 선택된 매체 리스트 출력
        print("\n✅ 현재 선택된 매체:")
        for i, media in enumerate(selected_media, 1):
            print(f"  {i}. {media['name']} ({media['category']})")
        
        # 추가 선택 여부 확인
        if max_count and len(selected_media) >= max_count:
            print(f"\n✅ 최대 {max_count}개 매체 선택 완료!")
            break
        
        while True:
            add_more = input("\n추가 매체 선택? (y/n): ").strip().lower()
            if add_more in ['y', 'n']:
                break
            print("❌ 'y' 또는 'n'을 입력해주세요.")
        
        if add_more == 'n':
            break
    
    return selected_media


def input_media_details(selected_media):
    """
    각 매체별 상세 데이터 입력 함수
    
    Args:
        selected_media: 선택된 매체 리스트
    
    Returns:
        상세 데이터가 추가된 매체 리스트
    """
    if not selected_media:
        return selected_media
    
    while True:
        total_ratio = 0
        
        print("\n" + "="*50)
        print("📝 매체별 상세 데이터 입력")
        print("="*50)
        
        for i, media in enumerate(selected_media, 1):
            print(f"\n{'─'*50}")
            print(f"[{i}/{len(selected_media)}] {media['name']} ({media['category']})")
            print('─'*50)
            
            # 1. 예산 비중 입력 (%)
            budget_ratio = get_number_input("예산 비중 (%): ", allow_zero=True, allow_negative=False, min_val=0, max_val=100)
            media['budget_ratio'] = budget_ratio
            total_ratio += budget_ratio
            print(f"📊 현재까지 입력된 비중 합계: {total_ratio:.2f}%")
            
            # 2. CPC 입력 (원)
            cpc = get_number_input("CPC (원): ", allow_zero=False, allow_negative=False)
            media['cpc'] = int(cpc)
            
            # 3. CTR 입력 (%)
            ctr = get_number_input("CTR (%): ", allow_zero=False, allow_negative=False)
            media['ctr'] = round(ctr, 2)
            
            # 4. CVR 입력 (%)
            cvr = get_number_input("CVR (%): ", allow_zero=False, allow_negative=False)
            media['cvr'] = round(cvr, 2)
            
            # 5. 전환당 매출액 입력 (원)
            revenue_per_cv = get_number_input("전환당 매출액 (원): ", allow_zero=False, allow_negative=False)
            media['revenue_per_cv'] = int(revenue_per_cv)
            
            # 6. 예측 오차 범위 입력 (%)
            adjustment = get_number_input("예측 오차 범위 (%) [-50 ~ +50]: ", allow_zero=True, allow_negative=True, min_val=-50, max_val=50)
            media['adjustment'] = adjustment
        
        # 예산 비중 합계 검증
        print("\n" + "="*50)
        print(f"💯 예산 비중 합계: {total_ratio:.2f}%")
        
        if abs(total_ratio - 100) < 0.01:  # 부동소수점 오차 고려
            print("✅ 예산 비중 합계가 100%입니다!")
            break
        else:
            print(f"❌ 예산 비중 합계: {total_ratio:.2f}%. 100%가 되어야 합니다.")
            
            while True:
                modify_choice = input("\n수정 방법을 선택하세요:\n  1. 개별 매체 수정\n  2. 전체 재입력\n  3. 무시하고 계속\n선택 (1-3): ").strip()
                
                if modify_choice == '1':
                    # 개별 매체 수정
                    modified_media = modify_budget_ratio(selected_media)
                    if modified_media is None:
                        # 전체 재입력 선택됨
                        break
                    else:
                        selected_media = modified_media
                        total_ratio = sum(m['budget_ratio'] for m in selected_media)
                        
                        if abs(total_ratio - 100) < 0.01:
                            print("\n✅ 예산 비중 합계가 100%입니다!")
                            break
                        else:
                            print(f"\n⚠️ 여전히 합계가 {total_ratio:.2f}%입니다.")
                            continue
                
                elif modify_choice == '2':
                    # 전체 재입력
                    break
                
                elif modify_choice == '3':
                    # 무시하고 계속
                    print("⚠️ 예산 비중 합계가 100%가 아닙니다. 계산 결과가 정확하지 않을 수 있습니다.")
                    break
                
                else:
                    print("❌ 1, 2, 또는 3을 입력해주세요.")
            
            if modify_choice in ['1', '3'] and abs(total_ratio - 100) < 0.01:
                break
            elif modify_choice == '3':
                break
    
    # 입력 완료된 데이터 테이블 형태로 출력
    print("\n" + "="*50)
    print("📊 입력 데이터 요약")
    print("="*50)
    
    print(f"\n{'매체명':<15} {'카테고리':<12} {'비중':<8} {'CPC':<10} {'CTR':<8} {'CVR':<8} {'전환매출':<12} {'오차범위'}")
    print("-"*100)
    
    for media in selected_media:
        print(f"{media['name']:<15} "
              f"{media['category']:<12} "
              f"{media['budget_ratio']:>6.2f}% "
              f"{format_number(media['cpc']):>10}원 "
              f"{media['ctr']:>6.2f}% "
              f"{media['cvr']:>6.2f}% "
              f"{format_number(media['revenue_per_cv']):>12}원 "
              f"{media['adjustment']:>+6.1f}%")
    
    return selected_media


def get_scenario_adjustment():
    """
    시나리오 조정 폭 입력 함수
    
    Returns:
        조정 폭 (%)
    """
    print("\n" + "="*50)
    print("🎯 시나리오 조정 폭 설정")
    print("="*50)
    print("시나리오 조정 폭을 입력하세요 (디폴트: ±5%)")
    print("보수안: -X%, 기본안: 0%, 공격안: +X%")
    
    while True:
        adjustment_input = input("\n조정 폭 (1~30, 엔터=5): ").strip()
        
        # 엔터 입력 시 기본값 5% 적용
        if adjustment_input == "":
            adjustment = 5.0
            print(f"✅ 조정 폭: ±{adjustment}% (기본값)")
            return adjustment
        
        try:
            adjustment_input = adjustment_input.replace(',', '')
            adjustment = float(adjustment_input)
            
            if 1 <= adjustment <= 30:
                print(f"✅ 조정 폭: ±{adjustment}%")
                return adjustment
            else:
                print("❌ 1~30 사이 값을 입력하세요.")
        except ValueError:
            print("❌ 숫자만 입력 가능합니다.")


def calculate_media_performance(media, total_budget):
    """
    매체별 성과 계산 함수
    
    Args:
        media: 매체 정보 딕셔너리
        total_budget: 총 예산
    
    Returns:
        성과 데이터가 추가된 매체 딕셔너리
    
    Raises:
        ValueError: 잘못된 입력 값
        ZeroDivisionError: 0으로 나누기 오류
    """
    try:
        # 입력값 검증
        if total_budget <= 0:
            raise ValueError(f"총 예산은 0보다 커야 합니다: {total_budget}")
        if media.get('budget_ratio', 0) < 0:
            raise ValueError(f"예산 비중은 0 이상이어야 합니다: {media.get('budget_ratio')}")
        if media.get('cpc', 0) <= 0:
            raise ValueError(f"CPC는 0보다 커야 합니다: {media.get('cpc')}")
        if media.get('ctr', 0) <= 0:
            raise ValueError(f"CTR은 0보다 커야 합니다: {media.get('ctr')}")
        
        # a) 매체 예산 = 총예산 × (예산비중 / 100)
        media_budget = total_budget * (media['budget_ratio'] / 100)
        
        # b) 예상 클릭수 = 매체예산 / CPC
        estimated_clicks = media_budget / media['cpc']
        
        # c) 예상 노출수 = 예상클릭수 / (CTR / 100)
        estimated_impressions = estimated_clicks / (media['ctr'] / 100)
        
        # d) CPM = (매체예산 / 예상노출수) × 1000
        cpm = (media_budget / estimated_impressions) * 1000 if estimated_impressions > 0 else 0
        
        # e) 예상 전환수 = 예상클릭수 × (CVR / 100)
        estimated_conversions = estimated_clicks * (media.get('cvr', 0) / 100)
        
        # f) CPA = 매체예산 / 예상전환수
        cpa = media_budget / estimated_conversions if estimated_conversions > 0 else 0
        
        # g) 총 매출 = 예상전환수 × 전환당매출액
        total_revenue = estimated_conversions * media.get('revenue_per_cv', 0)
        
        # h) ROAS = (총매출 / 매체예산) × 100
        roas = (total_revenue / media_budget) * 100 if media_budget > 0 else 0
        
        # i) 예측 오차 적용
        # 예상전환수 = 예상전환수 × (1 + 예측오차/100)
        estimated_conversions_adjusted = estimated_conversions * (1 + media.get('adjustment', 0) / 100)
        
        # CPA, ROAS 재계산
        cpa_adjusted = media_budget / estimated_conversions_adjusted if estimated_conversions_adjusted > 0 else 0
        total_revenue_adjusted = estimated_conversions_adjusted * media.get('revenue_per_cv', 0)
        roas_adjusted = (total_revenue_adjusted / media_budget) * 100 if media_budget > 0 else 0
        
        # 계산 결과 저장
        performance = {
            'media_budget': media_budget,
            'estimated_impressions': estimated_impressions,
            'estimated_clicks': estimated_clicks,
            'cpm': cpm,
            'estimated_conversions': estimated_conversions,
            'estimated_conversions_adjusted': estimated_conversions_adjusted,
            'cpa': cpa,
            'cpa_adjusted': cpa_adjusted,
            'total_revenue': total_revenue,
            'total_revenue_adjusted': total_revenue_adjusted,
            'roas': roas,
            'roas_adjusted': roas_adjusted
        }
        
        # 원본 매체 정보와 성과 데이터 병합
        return {**media, **performance}
    
    except ZeroDivisionError as e:
        raise ZeroDivisionError(f"매체 '{media.get('name', '알수없음')}' 계산 중 0으로 나누기 오류: {e}")
    except KeyError as e:
        raise KeyError(f"매체 '{media.get('name', '알수없음')}' 필수 데이터 누락: {e}")
    except Exception as e:
        raise Exception(f"매체 '{media.get('name', '알수없음')}' 성과 계산 중 오류: {e}")


def generate_scenarios(selected_media, total_budget, scenario_adjustment):
    """
    3개 시나리오 생성 함수 (안전성 강화)
    
    Args:
        selected_media: 선택된 매체 리스트
        total_budget: 총 예산
        scenario_adjustment: 시나리오 조정 폭 (%)
    
    Returns:
        시나리오 딕셔너리 {'conservative': [], 'base': [], 'aggressive': []}
    """
    scenarios = {
        'conservative': [],  # 보수안
        'base': [],          # 기본안
        'aggressive': []     # 공격안
    }
    
    for media in selected_media:
        try:
            # 기본 성과 계산
            media_performance = calculate_media_performance(media, total_budget)
            
            # 기본안 (예측오차만 적용된 값)
            base_media = media_performance.copy()
            scenarios['base'].append(base_media)
            
            # 보수안 (예상전환수에 추가로 -조정폭% 적용)
            conservative_media = media_performance.copy()
            conservative_media['estimated_conversions_adjusted'] = (
                media_performance['estimated_conversions_adjusted'] * (1 - scenario_adjustment / 100)
            )
            # CPA, ROAS 재계산
            conservative_media['cpa_adjusted'] = (
                conservative_media['media_budget'] / conservative_media['estimated_conversions_adjusted']
                if conservative_media['estimated_conversions_adjusted'] > 0 else 0
            )
            conservative_media['total_revenue_adjusted'] = (
                conservative_media['estimated_conversions_adjusted'] * conservative_media['revenue_per_cv']
            )
            conservative_media['roas_adjusted'] = (
                (conservative_media['total_revenue_adjusted'] / conservative_media['media_budget']) * 100
                if conservative_media['media_budget'] > 0 else 0
            )
            scenarios['conservative'].append(conservative_media)
            
            # 공격안 (예상전환수에 추가로 +조정폭% 적용)
            aggressive_media = media_performance.copy()
            aggressive_media['estimated_conversions_adjusted'] = (
                media_performance['estimated_conversions_adjusted'] * (1 + scenario_adjustment / 100)
            )
            # CPA, ROAS 재계산
            aggressive_media['cpa_adjusted'] = (
                aggressive_media['media_budget'] / aggressive_media['estimated_conversions_adjusted']
                if aggressive_media['estimated_conversions_adjusted'] > 0 else 0
            )
            aggressive_media['total_revenue_adjusted'] = (
                aggressive_media['estimated_conversions_adjusted'] * aggressive_media['revenue_per_cv']
            )
            aggressive_media['roas_adjusted'] = (
                (aggressive_media['total_revenue_adjusted'] / aggressive_media['media_budget']) * 100
                if aggressive_media['media_budget'] > 0 else 0
            )
            scenarios['aggressive'].append(aggressive_media)
        
        except Exception as e:
            # 매체별 계산 실패 시 안전 처리
            media_name = media.get('name', 'Unknown')
            print(f"⚠️ {media_name} 계산 실패: {str(e)}")
            
            # 계산 불가 표시용 더미 데이터
            error_media = {
                'name': media_name,
                'category': media.get('category', 'N/A'),
                'budget_ratio': media.get('budget_ratio', 0),
                'media_budget': 0,
                'cpm': 0,
                'cpc': media.get('cpc', 0),
                'ctr': media.get('ctr', 0),
                'cvr': media.get('cvr', 0),
                'estimated_impressions': 0,
                'estimated_clicks': 0,
                'estimated_conversions': 0,
                'estimated_conversions_adjusted': 0,
                'cpa_adjusted': 0,
                'revenue_per_cv': media.get('revenue_per_cv', 0),
                'total_revenue_adjusted': 0,
                'roas_adjusted': 0,
                'error': True,
                'error_message': f'계산 실패: {str(e)}'
            }
            
            scenarios['base'].append(error_media.copy())
            scenarios['conservative'].append(error_media.copy())
            scenarios['aggressive'].append(error_media.copy())
    
    return scenarios


def quick_mode():
    """빠른 모드 - 매체 3개 이하"""
    print("\n" + "="*50)
    print("🚀 빠른 모드 (매체 3개 이하)")
    print("="*50)
    
    # 1. 총 예산 입력
    print_progress(1, 5, "총 예산 입력")
    budget = get_budget_input()
    
    # 2. 매체 선택 (최대 3개)
    print_progress(2, 5, "매체 선택")
    selected_media = select_media(max_count=3)
    
    if not selected_media:
        print("\n⚠️ 선택된 매체가 없습니다.")
        input("\n계속하려면 Enter를 누르세요...")
        return
    
    print("\n✅ 매체 선택 완료")
    
    # 3. 매체별 상세 데이터 입력
    print_progress(3, 5, "매체별 상세 데이터 입력")
    selected_media = input_media_details(selected_media)
    
    # 3-1. 입력값 확인
    while True:
        confirm_result = confirm_input_data(budget, selected_media)
        
        if confirm_result == 'continue':
            break
        elif confirm_result == 'modify':
            # 수정 기능
            modified_media = modify_budget_ratio(selected_media)
            if modified_media:
                selected_media = modified_media
        elif confirm_result == 'save':
            # 중간 저장
            save_intermediate_data(budget, selected_media)
            break
    
    # 4. 시나리오 조정 폭 입력
    print_progress(4, 5, "시나리오 조정 폭 설정")
    scenario_adjustment = get_scenario_adjustment()
    
    # 5. 시나리오 생성 및 계산
    print_progress(5, 5, "성과 계산 및 결과 저장")
    print("\n⏳ 성과 계산 중...")
    scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
    print("✓ 3개 시나리오 계산 완료")
    
    # 시나리오 요약 미리보기 출력
    print_summary_preview(scenarios, budget)
    
    # Excel 파일 저장
    print("\n⏳ Excel 파일 저장 중...")
    filename = save_to_excel(scenarios, budget, scenario_adjustment)
    
    # 전체 경로 출력
    full_path = os.path.abspath(filename)
    print(f"✓ 파일 저장: {full_path}")
    
    # 최종 완료 메시지
    print("\n" + "="*50)
    print("✅ 빠른 모드 완료!")
    print("="*50)
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🎯 시나리오: 보수안(-{scenario_adjustment}%), 기본안(0%), 공격안(+{scenario_adjustment}%)")
    print(f"📁 저장 파일: {filename}")
    
    # 10. 계속 진행 메뉴
    while True:
        print("\n계속 진행하시겠습니까?")
        print("  1. 새로 시작")
        print("  2. 종료")
        continue_choice = get_number_input("선택 (1-2): ", allow_zero=False, allow_negative=False)
        
        if continue_choice == 1:
            return  # 메인 메뉴로 돌아감
        elif continue_choice == 2:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            exit()
        else:
            print("\n❌ 1 또는 2를 입력해주세요.")


def detailed_mode():
    """상세 모드"""
    print("\n" + "="*50)
    print("📊 상세 모드")
    print("="*50)
    
    # 1. 총 예산 입력
    print_progress(1, 5, "총 예산 입력")
    budget = get_budget_input()
    
    # 2. 매체 선택 (제한 없음)
    print_progress(2, 5, "매체 선택")
    selected_media = select_media(max_count=None)
    
    if not selected_media:
        print("\n⚠️ 선택된 매체가 없습니다.")
        input("\n계속하려면 Enter를 누르세요...")
        return
    
    print("\n✅ 매체 선택 완료")
    
    # 3. 매체별 상세 데이터 입력
    print_progress(3, 5, "매체별 상세 데이터 입력")
    selected_media = input_media_details(selected_media)
    
    # 3-1. 입력값 확인
    while True:
        confirm_result = confirm_input_data(budget, selected_media)
        
        if confirm_result == 'continue':
            break
        elif confirm_result == 'modify':
            # 수정 기능
            modified_media = modify_budget_ratio(selected_media)
            if modified_media:
                selected_media = modified_media
        elif confirm_result == 'save':
            # 중간 저장
            save_intermediate_data(budget, selected_media)
            break
    
    # 4. 시나리오 조정 폭 입력
    print_progress(4, 5, "시나리오 조정 폭 설정")
    scenario_adjustment = get_scenario_adjustment()
    
    # 5. 시나리오 생성 및 계산
    print_progress(5, 5, "성과 계산 및 결과 저장")
    print("\n⏳ 성과 계산 중...")
    scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
    print("✓ 3개 시나리오 계산 완료")
    
    # 시나리오 요약 미리보기 출력
    print_summary_preview(scenarios, budget)
    
    # Excel 파일 저장
    print("\n⏳ Excel 파일 저장 중...")
    filename = save_to_excel(scenarios, budget, scenario_adjustment)
    
    # 전체 경로 출력
    full_path = os.path.abspath(filename)
    print(f"✓ 파일 저장: {full_path}")
    
    # 최종 완료 메시지
    print("\n" + "="*50)
    print("✅ 상세 모드 완료!")
    print("="*50)
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🎯 시나리오: 보수안(-{scenario_adjustment}%), 기본안(0%), 공격안(+{scenario_adjustment}%)")
    print(f"📁 저장 파일: {filename}")
    
    # 10. 계속 진행 메뉴
    while True:
        print("\n계속 진행하시겠습니까?")
        print("  1. 새로 시작")
        print("  2. 종료")
        continue_choice = get_number_input("선택 (1-2): ", allow_zero=False, allow_negative=False)
        
        if continue_choice == 1:
            return  # 메인 메뉴로 돌아감
        elif continue_choice == 2:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            exit()
        else:
            print("\n❌ 1 또는 2를 입력해주세요.")


def create_scenario_dataframe(scenario_data, total_budget):
    """
    시나리오 데이터를 DataFrame으로 변환
    
    Args:
        scenario_data: 시나리오 매체 리스트
        total_budget: 총 예산
    
    Returns:
        pandas DataFrame
    """
    data = []
    
    for media in scenario_data:
        row = {
            '매체명': media['name'],
            '카테고리': media['category'],
            '예산': int(media['media_budget']),
            '예산비중': media['budget_ratio'],
            'CPM': int(media['cpm']),
            '예상노출': int(media['estimated_impressions']),
            '예상클릭': int(media['estimated_clicks']),
            'CTR': media['ctr'],
            'CPC': media['cpc'],
            '예상전환': round(media['estimated_conversions_adjusted'], 1),
            'CVR': media['cvr'],
            'CPA': int(media['cpa_adjusted']),
            'ROAS': round(media['roas_adjusted'], 1)
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    
    # 합계 행 추가
    total_budget_sum = df['예산'].sum()
    total_impressions = df['예상노출'].sum()
    total_clicks = df['예상클릭'].sum()
    total_conversions = df['예상전환'].sum()
    
    # 가중평균 계산
    avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    avg_cvr = (total_conversions / total_clicks * 100) if total_clicks > 0 else 0
    avg_cpa = (total_budget_sum / total_conversions) if total_conversions > 0 else 0
    
    # ROAS 가중평균
    total_revenue = sum(media['total_revenue_adjusted'] for media in scenario_data)
    avg_roas = (total_revenue / total_budget_sum * 100) if total_budget_sum > 0 else 0
    
    total_row = {
        '매체명': '합계',
        '카테고리': '',
        '예산': int(total_budget_sum),
        '예산비중': 100.0,
        'CPM': '',
        '예상노출': int(total_impressions),
        '예상클릭': int(total_clicks),
        'CTR': round(avg_ctr, 2),
        'CPC': '',
        '예상전환': round(total_conversions, 1),
        'CVR': round(avg_cvr, 2),
        'CPA': int(avg_cpa),
        'ROAS': round(avg_roas, 1)
    }
    
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    
    return df


def create_summary_dataframe(scenarios, total_budget):
    """
    시나리오 요약 DataFrame 생성
    
    Args:
        scenarios: 시나리오 딕셔너리
        total_budget: 총 예산
    
    Returns:
        pandas DataFrame
    """
    summary_data = []
    
    scenario_names = {
        'conservative': '보수안',
        'base': '기본안',
        'aggressive': '공격안'
    }
    
    for scenario_key, scenario_name in scenario_names.items():
        scenario_data = scenarios[scenario_key]
        
        total_conversions = sum(media['estimated_conversions_adjusted'] for media in scenario_data)
        total_clicks = sum(media['estimated_clicks'] for media in scenario_data)
        
        # 평균 CPA 계산
        avg_cpa = (total_budget / total_conversions) if total_conversions > 0 else 0
        
        # 평균 ROAS 계산
        total_revenue = sum(media['total_revenue_adjusted'] for media in scenario_data)
        avg_roas = (total_revenue / total_budget * 100) if total_budget > 0 else 0
        
        row = {
            '구분': scenario_name,
            '총전환수': int(total_conversions),
            '평균CPA': int(avg_cpa),
            '총클릭수': int(total_clicks),
            '평균ROAS': round(avg_roas, 1)
        }
        summary_data.append(row)
    
    df = pd.DataFrame(summary_data)
    return df


def adjust_column_width(worksheet, df):
    """
    컬럼 너비를 15로 고정
    
    Args:
        worksheet: openpyxl 워크시트
        df: pandas DataFrame
    """
    for idx, col in enumerate(df.columns, 1):
        # 모든 열 너비 15로 고정
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = 15


def apply_borders(worksheet, start_row, end_row, start_col, end_col):
    """
    테두리 적용
    
    Args:
        worksheet: openpyxl 워크시트
        start_row: 시작 행
        end_row: 끝 행
        start_col: 시작 열
        end_col: 끝 열
    """
    # 얇은 실선 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 모든 셀에 테두리 적용
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = thin_border


def save_to_excel(scenarios, total_budget, scenario_adjustment):
    """
    결과를 Excel 파일로 저장 (최적화 버전)
    
    최적화 내용:
    - 테두리: 헤더 + 합계행만 적용 (데이터 행 생략)
    - 퍼센트 변환: DataFrame 단계에서 일괄 처리
    - 열 너비: 한 번에 설정
    - 셀 루프 최소화
    
    Args:
        scenarios: 시나리오 딕셔너리
        total_budget: 총 예산
        scenario_adjustment: 시나리오 조정 폭
    
    Returns:
        저장된 파일명
    """
    # 파일명 생성: 미디어믹스_YYYY년M월_YYYYMMDD_HHMMSS.xlsx
    now = datetime.now()
    year = now.year
    month = now.month
    date_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"미디어믹스_{year}년{month}월_{date_str}.xlsx"
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Excel Writer 생성
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 1. 시나리오별 시트 생성
        scenario_sheet_names = {
            'conservative': f'보수안(-{scenario_adjustment}%)',
            'base': '기본안',
            'aggressive': f'공격안(+{scenario_adjustment}%)'
        }
        
        for scenario_key, sheet_name in scenario_sheet_names.items():
            df = create_scenario_dataframe(scenarios[scenario_key], total_budget)
            
            # 퍼센티지 컬럼을 DataFrame 단계에서 변환
            percentage_columns = ['예산비중', 'CTR', 'CVR', 'ROAS']
            for col in percentage_columns:
                if col in df.columns:
                    df[col] = df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) and x != '' else x)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 워크시트 포맷 적용
            worksheet = writer.sheets[sheet_name]
            num_rows = len(df) + 1  # 헤더 포함
            num_cols = len(df.columns)
            
            # 모든 열 너비 15로 설정
            for col_idx in range(1, num_cols + 1):
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[col_letter].width = 15
            
            # 가운데 정렬 (모든 셀)
            center_alignment = Alignment(horizontal='center', vertical='center')
            for row in range(1, num_rows + 1):
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=row, column=col).alignment = center_alignment
            
            # 테두리 적용: 헤더(row=1) + 합계행(마지막 row)만
            for col in range(1, num_cols + 1):
                # 헤더 행
                worksheet.cell(row=1, column=col).border = thin_border
                # 합계 행 (마지막 행)
                worksheet.cell(row=num_rows, column=col).border = thin_border
        
        # 2. 시나리오 요약 시트 생성
        summary_df = create_summary_dataframe(scenarios, total_budget)
        
        # 퍼센티지 컬럼 변환
        if '평균ROAS' in summary_df.columns:
            summary_df['평균ROAS'] = summary_df['평균ROAS'].apply(
                lambda x: f"{x:.1f}%" if pd.notna(x) and x != '' else x
            )
        
        summary_df.to_excel(writer, sheet_name='시나리오_요약', index=False)
        
        # 요약 시트 서식 설정
        worksheet = writer.sheets['시나리오_요약']
        num_rows = len(summary_df) + 1
        num_cols = len(summary_df.columns)
        
        # 모든 열 너비 15로 설정
        for col_idx in range(1, num_cols + 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = 15
        
        # 가운데 정렬
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                worksheet.cell(row=row, column=col).alignment = center_alignment
        
        # 테두리: 헤더 + 마지막 행만
        for col in range(1, num_cols + 1):
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=num_rows, column=col).border = thin_border
    
    return filename


def print_summary_preview(scenarios, total_budget):
    """
    터미널에 시나리오 요약 테이블 미리보기 출력
    
    Args:
        scenarios: 시나리오 딕셔너리
        total_budget: 총 예산
    """
    print("\n" + "="*50)
    print("📊 시나리오 요약")
    print("="*50)
    
    summary_df = create_summary_dataframe(scenarios, total_budget)
    
    print(f"\n{'구분':<12} {'총전환수':<15} {'평균CPA':<15} {'총클릭수':<15} {'평균ROAS'}")
    print("-"*80)
    
    for _, row in summary_df.iterrows():
        print(f"{row['구분']:<12} "
              f"{format_number(row['총전환수']):>13}건 "
              f"{format_number(row['평균CPA']):>13}원 "
              f"{format_number(row['총클릭수']):>13}회 "
              f"{row['평균ROAS']:>10.1f}%")


def create_excel_template():
    """
    Excel 입력 템플릿 생성
    """
    filename = "미디어믹스_입력.xlsx"
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # A1: "총예산", B1: 100000000
    ws['A1'] = "총예산"
    ws['B1'] = 100000000
    
    # 헤더 (A3~G3)
    headers = ["매체명", "카테고리", "예산비중", "평균 CPC", "평균 CTR", "평균 CVR", "예측 오차"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 예시 데이터
    sample_data = [
        ["네이버", "SA", "25%", 300, "3%", "4%", "1%"],
        ["구글", "SA", "25%", 400, "3%", "4%", "1%"],
        ["메타", "DA", "25%", 500, "3%", "4%", "10%"],
        ["카카오DA", "DA", "25%", 600, "3%", "4%", "-1%"]
    ]
    
    for row_idx, data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 모든 열 너비: 15
    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    # 파일 저장
    wb.save(filename)
    print(f"✓ Excel 템플릿 생성 완료: {filename}")
    print("\n템플릿 파일을 수정한 후 '6. Excel 파일에서 불러오기'를 선택하세요.")


def read_excel_input():
    """
    Excel 파일에서 데이터 읽기
    
    Returns:
        (budget, selected_media) 튜플 또는 None
    """
    filename = "미디어믹스_입력.xlsx"
    
    if not os.path.exists(filename):
        print(f"\n❌ '{filename}' 파일을 찾을 수 없습니다.")
        print("먼저 '5. Excel 템플릿 생성'을 실행하세요.")
        input("\n계속하려면 Enter를 누르세요...")
        return None
    
    try:
        # Excel 파일 읽기
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        
        # B1셀에서 총예산 읽기
        budget = ws['B1'].value
        if budget is None or budget <= 0:
            print("\n❌ 총예산이 올바르지 않습니다. (B1 셀 확인)")
            input("\n계속하려면 Enter를 누르세요...")
            return None
        
        # 카테고리 매핑
        category_map = {
            "SA": "검색광고",
            "DA": "디스플레이광고",
            "VA": "동영상광고"
        }
        
        # A4행부터 데이터 읽기
        selected_media = []
        row = 4
        total_ratio = 0
        
        while True:
            media_name = ws.cell(row=row, column=1).value
            
            # A열이 빈 셀이면 종료
            if media_name is None or str(media_name).strip() == "":
                break
            
            category_code = ws.cell(row=row, column=2).value
            budget_ratio_raw = ws.cell(row=row, column=3).value
            cpc = ws.cell(row=row, column=4).value
            ctr_raw = ws.cell(row=row, column=5).value
            cvr_raw = ws.cell(row=row, column=6).value
            adjustment_raw = ws.cell(row=row, column=7).value
            
            # % 문자 제거 및 숫자 변환
            def parse_percentage(value):
                if isinstance(value, str):
                    value = value.replace('%', '').strip()
                try:
                    return float(value)
                except:
                    return 0
            
            budget_ratio = parse_percentage(budget_ratio_raw)
            ctr = parse_percentage(ctr_raw)
            cvr = parse_percentage(cvr_raw)
            adjustment = parse_percentage(adjustment_raw)
            
            # 카테고리 변환
            category = category_map.get(str(category_code).upper(), "검색광고")
            
            # 매체 데이터 추가
            media = {
                'name': str(media_name).strip(),
                'category': category,
                'budget_ratio': budget_ratio,
                'cpc': int(cpc) if cpc else 0,
                'ctr': ctr,
                'cvr': cvr,
                'revenue_per_cv': 100000,  # 기본값 (Excel에서 입력받지 않음)
                'adjustment': adjustment
            }
            
            selected_media.append(media)
            total_ratio += budget_ratio
            row += 1
        
        if not selected_media:
            print("\n❌ 읽을 수 있는 매체 데이터가 없습니다.")
            input("\n계속하려면 Enter를 누르세요...")
            return None
        
        # 불러온 데이터 요약 출력
        print("\n" + "="*50)
        print("✅ Excel 데이터 로드 완료")
        print("="*50)
        print(f"\n💰 총 예산: {format_number(budget)}원")
        print(f"📱 매체 수: {len(selected_media)}개")
        print(f"💯 예산 비중 합계: {total_ratio:.2f}%")
        
        # 예산 비중 합계 검증
        if abs(total_ratio - 100) > 0.01:
            print(f"\n⚠️ 경고: 예산 비중 합계가 {total_ratio:.2f}%입니다. 100%가 되어야 합니다.")
            print("Excel 파일을 수정하거나 계속 진행하세요.")
        else:
            print("\n✅ 예산 비중 합계가 100%입니다!")
        
        print("\n📋 매체 목록:")
        for i, media in enumerate(selected_media, 1):
            print(f"  {i}. {media['name']} ({media['category']}) - 비중: {media['budget_ratio']:.2f}%")
        
        # 전환당 매출액 입력
        print("\n" + "="*50)
        print("📝 전환당 매출액 입력")
        print("="*50)
        print("(Excel에는 없는 항목으로 추가 입력이 필요합니다)")
        
        for media in selected_media:
            revenue = get_number_input(f"\n{media['name']} - 전환당 매출액 (원): ", 
                                      allow_zero=False, allow_negative=False)
            media['revenue_per_cv'] = int(revenue)
        
        # 이어서 진행 여부 확인
        print("\n" + "-"*50)
        while True:
            continue_choice = input("이어서 진행하시겠습니까? (y/n): ").strip().lower()
            if continue_choice in ['y', 'n']:
                break
            print("❌ 'y' 또는 'n'을 입력해주세요.")
        
        if continue_choice == 'y':
            return (budget, selected_media)
        else:
            return None
    
    except Exception as e:
        print(f"\n❌ Excel 파일 읽기 오류: {e}")
        input("\n계속하려면 Enter를 누르세요...")
        return None


def apply_budget_adjustment(cpc, budget):
    """
    예산 규모에 따른 CPC 보정
    
    Args:
        cpc: 기본 CPC
        budget: 총 예산
    
    Returns:
        보정된 CPC
    """
    if budget <= 10000000:  # 1천만원 이하
        return int(cpc * 0.9)
    elif budget <= 50000000:  # 1천만~5천만
        return int(cpc)
    elif budget <= 100000000:  # 5천만~1억
        return int(cpc * 1.1)
    else:  # 1억 이상
        return int(cpc * 1.2)


def ai_prediction_mode():
    """AI 자동 예측 모드"""
    print("\n" + "="*50)
    print("🤖 AI 자동 예측 모드")
    print("="*50)
    
    # 1. 총 예산 입력
    print("\n💰 총 예산을 입력하세요")
    budget = get_number_input("예산 (원): ", allow_zero=False, allow_negative=False)
    print(f"✅ 입력된 예산: {format_number(budget)}원")
    
    # 2. 업종 선택
    print("\n" + "="*50)
    print("📋 업종을 선택하세요")
    print("="*50)
    industries = list(INDUSTRY_BASE_METRICS.keys())  # 새로운 Base Metrics 사용
    for i, industry in enumerate(industries, 1):
        print(f"  {i}. {industry}")
    
    while True:
        industry_choice = int(get_number_input(f"\n업종 선택 (1-{len(industries)}): ", 
                                               allow_zero=False, allow_negative=False, 
                                               min_val=1, max_val=len(industries)))
        if 1 <= industry_choice <= len(industries):
            selected_industry = industries[industry_choice - 1]
            print(f"✅ 선택된 업종: {selected_industry}")
            break
    
    # 3. 운영 월 입력
    print("\n📅 운영 월을 입력하세요 (1~12)")
    month = int(get_number_input("월: ", allow_zero=False, allow_negative=False, min_val=1, max_val=12))
    seasonality_factor = SEASONALITY_COMMON[month]  # 새로운 계절성 보정 사용
    
    # 업종별 계절성 추가 정보 표시
    if selected_industry in INDUSTRY_SEASON_WEIGHT:
        season_info = INDUSTRY_SEASON_WEIGHT[selected_industry]
        if month in season_info['high_months']:
            print(f"✅ {month}월 계절성: {seasonality_factor:.2f}x (업종 성수기 +{(season_info['high_multiplier']-1)*100:.0f}%)")
        elif month in season_info['low_months']:
            print(f"✅ {month}월 계절성: {seasonality_factor:.2f}x (업종 비수기 {(season_info['low_multiplier']-1)*100:.0f}%)")
        else:
            print(f"✅ {month}월 계절성: {seasonality_factor:.2f}x")
    else:
        print(f"✅ {month}월 계절성: {seasonality_factor:.2f}x")
    
    # 4. 목표 선택
    print("\n" + "="*50)
    print("🎯 마케팅 목표를 선택하세요")
    print("="*50)
    print("  1. 인지도 (DA 중심)")
    print("  2. 전환 (SA 중심)")
    print("  3. 균형 (SA/DA 균형)")
    
    goal_choice = int(get_number_input("\n목표 선택 (1-3): ", allow_zero=False, allow_negative=False, min_val=1, max_val=3))
    
    if goal_choice == 1:
        goal_name = "인지도"
        da_ratio = 0.6
        sa_ratio = 0.4
    elif goal_choice == 2:
        goal_name = "전환"
        da_ratio = 0.3
        sa_ratio = 0.7
    else:
        goal_name = "균형"
        da_ratio = 0.5
        sa_ratio = 0.5
    
    print(f"✅ 선택된 목표: {goal_name} (SA: {sa_ratio*100:.0f}%, DA: {da_ratio*100:.0f}%)")
    
    # 5. 제외할 매체 선택 (선택사항)
    all_media = list(MEDIA_MULTIPLIERS.keys())  # 모든 매체 목록
    print("\n" + "="*50)
    print("🚫 제외할 매체를 선택하세요 (선택사항)")
    print("="*50)
    for i, media in enumerate(all_media, 1):
        media_name = media.split('_')[0]
        media_type = media.split('_')[1] if '_' in media else ''
        type_label = {'SA': '검색', 'DA': '디스플레이', 'VA': '동영상'}.get(media_type, '')
        print(f"  {i}. {media_name} ({type_label})")
    print("  0. 모든 매체 사용")
    
    excluded = []
    while True:
        exclude_choice = int(get_number_input(f"\n제외할 매체 번호 (0-{len(all_media)}, 0=사용 안 함): ", 
                                              allow_zero=True, allow_negative=False, 
                                              min_val=0, max_val=len(all_media)))
        if exclude_choice == 0:
            break
        excluded.append(all_media[exclude_choice - 1])
        print(f"✅ {all_media[exclude_choice - 1].split('_')[0]} 제외됨")
        
        more = input("더 제외할 매체가 있습니까? (y/n): ").strip().lower()
        if more != 'y':
            break
    
    # 6. AI 자동 매체 선택 및 예산 배분 (새로운 보정 시스템 적용)
    print("\n" + "="*50)
    print("🤖 AI가 최적 미디어믹스를 생성 중...")
    print("="*50)
    print(f"   - 업종별 Base Metrics 로드: {selected_industry}")
    print(f"   - 계절성 보정 적용: {month}월")
    print(f"   - 예산 규모 경쟁도 보정: {format_number(budget)}원")
    
    selected_media = []
    
    # SA 매체 필터링 (검색광고)
    sa_media = [k for k in MEDIA_MULTIPLIERS.keys() 
                if '_SA' in k and k not in excluded]
    # DA 매체 필터링 (디스플레이광고)
    da_media = [k for k in MEDIA_MULTIPLIERS.keys() 
                if '_DA' in k and k not in excluded]
    # VA 매체 필터링 (동영상광고)
    va_media = [k for k in MEDIA_MULTIPLIERS.keys() 
                if '_VA' in k and k not in excluded]
    
    sa_budget = budget * sa_ratio
    da_budget = budget * da_ratio
    
    # SA 매체 예산 배분
    if sa_media:
        sa_budget_per_media = sa_budget / len(sa_media)
        for media_key in sa_media:
            media_name = media_key.split('_')[0]
            
            # 새로운 보정 시스템 적용: 업종 Base + 매체 Multiplier + 계절성 + 예산 경쟁도
            adjusted_metrics = get_media_adjusted_metrics(
                selected_industry, 
                media_key, 
                month, 
                sa_budget_per_media
            )
            
            if adjusted_metrics:
                media = {
                    'name': media_name,
                    'category': '검색광고',
                    'budget_ratio': (sa_budget_per_media / budget) * 100,
                    'cpc': round(adjusted_metrics['CPC'], 0),
                    'ctr': round(adjusted_metrics['CTR'] * 100, 2),  # % 변환
                    'cvr': round(adjusted_metrics['CVR'] * 100, 2),  # % 변환
                    'revenue_per_cv': 100000,  # 기본값
                    'adjustment': 0  # 기본안
                }
                selected_media.append(media)
    
    # DA 매체 예산 배분
    if da_media:
        da_budget_per_media = da_budget / len(da_media)
        for media_key in da_media:
            media_name = media_key.replace('_DA', '').replace('DA', '')
            
            # 새로운 보정 시스템 적용
            adjusted_metrics = get_media_adjusted_metrics(
                selected_industry, 
                media_key, 
                month, 
                da_budget_per_media
            )
            
            if adjusted_metrics:
                media = {
                    'name': media_name,
                    'category': '디스플레이광고',
                    'budget_ratio': (da_budget_per_media / budget) * 100,
                    'cpc': round(adjusted_metrics['CPC'], 0),
                    'ctr': round(adjusted_metrics['CTR'] * 100, 2),
                    'cvr': round(adjusted_metrics['CVR'] * 100, 2),
                    'revenue_per_cv': 80000,  # 기본값
                    'adjustment': 0  # 기본안
                }
                selected_media.append(media)
    
    # VA 매체 예산 배분 (있을 경우)
    if va_media and da_budget > 0:
        va_budget_per_media = da_budget * 0.2 / len(va_media)  # DA 예산의 20%
        for media_key in va_media:
            media_name = media_key.replace('_VA', '').replace('VA', '')
            
            adjusted_metrics = get_media_adjusted_metrics(
                selected_industry, 
                media_key, 
                month, 
                va_budget_per_media
            )
            
            if adjusted_metrics:
                media = {
                    'name': media_name,
                    'category': '동영상광고',
                    'budget_ratio': (va_budget_per_media / budget) * 100,
                    'cpc': round(adjusted_metrics['CPC'], 0),
                    'ctr': round(adjusted_metrics['CTR'] * 100, 2),
                    'cvr': round(adjusted_metrics['CVR'] * 100, 2),
                    'revenue_per_cv': 70000,  # 기본값
                    'adjustment': 0  # 기본안
                }
                selected_media.append(media)
    
    # 예산 비중 정규화 (합계 100%)
    total_ratio = sum(m['budget_ratio'] for m in selected_media)
    for media in selected_media:
        media['budget_ratio'] = (media['budget_ratio'] / total_ratio) * 100
    
    print(f"✓ AI가 {selected_industry} 업종 기준으로 최적 미디어믹스를 생성했습니다")
    
    # 7. 생성된 미디어믹스 출력
    print("\n" + "="*50)
    print("📊 AI 생성 미디어믹스")
    print("="*50)
    
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🎯 목표: {goal_name}")
    print(f"📅 운영 월: {month}월 (계절성: {seasonality_factor:.2f}x)")
    
    print(f"\n{'매체명':<15} {'카테고리':<12} {'비중':<8} {'CPC':<10} {'CTR':<8} {'CVR':<8}")
    print("-"*70)
    
    for media in selected_media:
        print(f"{media['name']:<15} "
              f"{media['category']:<12} "
              f"{media['budget_ratio']:>6.1f}% "
              f"{format_number(media['cpc']):>10}원 "
              f"{media['ctr']:>6.1f}% "
              f"{media['cvr']:>6.1f}%")
    
    # 8. 시나리오 조정 폭 (자동 설정)
    print("\n" + "="*50)
    print("🎯 시나리오 자동 설정")
    print("="*50)
    print("보수안: -5%, 기본안: 0%, 공격안: +10%")
    scenario_adjustment = 7.5  # 중간값
    
    # 시나리오별로 adjustment 설정
    scenarios_data = {
        'conservative': [],
        'base': [],
        'aggressive': []
    }
    
    for media in selected_media:
        # 보수안
        conservative_media = media.copy()
        conservative_media['adjustment'] = -5
        scenarios_data['conservative'].append(conservative_media)
        
        # 기본안
        base_media = media.copy()
        base_media['adjustment'] = 0
        scenarios_data['base'].append(base_media)
        
        # 공격안
        aggressive_media = media.copy()
        aggressive_media['adjustment'] = 10
        scenarios_data['aggressive'].append(aggressive_media)
    
    # 9. 시나리오 생성 및 계산
    print("\n⏳ 성과 계산 중...")
    scenarios = {}
    for scenario_key, media_list in scenarios_data.items():
        scenarios[scenario_key] = []
        for media in media_list:
            media_performance = calculate_media_performance(media, budget)
            scenarios[scenario_key].append(media_performance)
    
    print("✓ 3개 시나리오 계산 완료")
    
    # 10. 시나리오 요약 미리보기 출력
    print_summary_preview(scenarios, budget)
    
    # 11. Excel 파일 저장
    print("\n⏳ Excel 파일 저장 중...")
    filename = save_to_excel(scenarios, budget, scenario_adjustment)
    
    # 전체 경로 출력
    full_path = os.path.abspath(filename)
    print(f"✓ 파일 저장: {full_path}")
    
    # 12. 최종 완료 메시지
    print("\n" + "="*50)
    print("✅ AI 자동 예측 완료!")
    print("="*50)
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🏢 업종: {selected_industry}")
    print(f"🎯 목표: {goal_name}")
    print(f"📁 저장 파일: {filename}")
    
    # 13. 계속 진행 메뉴
    while True:
        print("\n계속 진행하시겠습니까?")
        print("  1. 새로 시작")
        print("  2. 종료")
        continue_choice = get_number_input("선택 (1-2): ", allow_zero=False, allow_negative=False)
        
        if continue_choice == 1:
            return  # 메인 메뉴로 돌아감
        elif continue_choice == 2:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            exit()
        else:
            print("\n❌ 1 또는 2를 입력해주세요.")


def excel_input_mode():
    """Excel 입력 모드"""
    print("\n" + "="*50)
    print("📊 Excel 입력 모드")
    print("="*50)
    
    # 데이터 불러오기
    result = read_excel_input()
    
    if result is None:
        print("\n메인 메뉴로 돌아갑니다.")
        input("\n계속하려면 Enter를 누르세요...")
        return
    
    budget, selected_media = result
    
    # 시나리오 조정 폭 입력
    scenario_adjustment = get_scenario_adjustment()
    
    # 시나리오 생성 및 계산
    print("\n⏳ 성과 계산 중...")
    scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
    print("✓ 3개 시나리오 계산 완료")
    
    # 시나리오 요약 미리보기 출력
    print_summary_preview(scenarios, budget)
    
    # Excel 파일 저장
    print("\n⏳ Excel 파일 저장 중...")
    filename = save_to_excel(scenarios, budget, scenario_adjustment)
    
    # 전체 경로 출력
    full_path = os.path.abspath(filename)
    print(f"✓ 파일 저장: {full_path}")
    
    # 최종 완료 메시지
    print("\n" + "="*50)
    print("✅ Excel 입력 모드 완료!")
    print("="*50)
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🎯 시나리오: 보수안(-{scenario_adjustment}%), 기본안(0%), 공격안(+{scenario_adjustment}%)")
    print(f"📁 저장 파일: {filename}")
    
    # 계속 진행 메뉴
    while True:
        print("\n계속 진행하시겠습니까?")
        print("  1. 새로 시작")
        print("  2. 종료")
        continue_choice = get_number_input("선택 (1-2): ", allow_zero=False, allow_negative=False)
        
        if continue_choice == 1:
            return  # 메인 메뉴로 돌아감
        elif continue_choice == 2:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            exit()
        else:
            print("\n❌ 1 또는 2를 입력해주세요.")


def load_saved_data():
    """저장된 데이터 불러오기 및 시뮬레이션 계속 진행"""
    print("\n" + "="*50)
    print("📂 저장된 데이터 불러오기")
    print("="*50)
    
    # 데이터 불러오기
    result = load_saved_data_file()
    
    if result is None:
        print("\n메인 메뉴로 돌아갑니다.")
        input("\n계속하려면 Enter를 누르세요...")
        return
    
    budget, selected_media = result
    
    # 이미 상세 데이터가 입력되어 있는지 확인
    has_details = all('budget_ratio' in media and 'cpc' in media for media in selected_media)
    
    if not has_details:
        # 상세 데이터 미입력 → 입력 단계로
        print("\n📝 매체별 상세 데이터를 입력해주세요.")
        selected_media = input_media_details(selected_media)
        
        # 중간 저장 옵션
        save_intermediate_data(budget, selected_media)
    
    # 시나리오 조정 폭 입력
    scenario_adjustment = get_scenario_adjustment()
    
    # 시나리오 생성 및 계산
    print("\n⏳ 성과 계산 중...")
    scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
    print("✓ 계산 완료")
    
    # 시나리오 요약 미리보기 출력
    print_summary_preview(scenarios, budget)
    
    # Excel 파일 저장
    print("\n⏳ Excel 파일 저장 중...")
    filename = save_to_excel(scenarios, budget, scenario_adjustment)
    print(f"✓ 결과 저장 완료: {filename}")
    
    # 최종 완료 메시지
    print("\n" + "="*50)
    print("✅ 불러오기 완료!")
    print("="*50)
    print(f"\n💰 총 예산: {format_number(budget)}원")
    print(f"📱 매체 수: {len(selected_media)}개")
    print(f"🎯 시나리오: 보수안(-{scenario_adjustment}%), 기본안(0%), 공격안(+{scenario_adjustment}%)")
    print(f"📁 저장 파일: {filename}")
    
    # 계속 진행 메뉴
    while True:
        print("\n계속 진행하시겠습니까?")
        print("  1. 새로 시작")
        print("  2. 종료")
        continue_choice = get_number_input("선택 (1-2): ", allow_zero=False, allow_negative=False)
        
        if continue_choice == 1:
            return  # 메인 메뉴로 돌아감
        elif continue_choice == 2:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            exit()
        else:
            print("\n❌ 1 또는 2를 입력해주세요.")


def main():
    """메인 함수 - 모드 선택"""
    # 첫 실행 시 README.txt 생성
    create_readme()
    
    while True:
        print("\n" + "="*50)
        print("미디어믹스 시뮬레이터")
        print("Media Mix Simulator")
        print("="*50)
        print("\n📋 모드를 선택하세요:")
        print("  1. 빠른 모드 (매체 3개 이하)")
        print("  2. 상세 모드")
        print("  3. 저장된 데이터 불러오기")
        print("  4. Excel 템플릿 생성")
        print("  5. Excel 파일에서 불러오기")
        print("  6. AI 자동 예측 모드 🤖")
        print("  7. 종료")
        print("-"*50)
        
        choice = get_number_input("선택 (1-7): ", allow_zero=False, allow_negative=False, min_val=1, max_val=7)
        
        if choice == 1:
            quick_mode()
        elif choice == 2:
            detailed_mode()
        elif choice == 3:
            load_saved_data()
        elif choice == 4:
            create_excel_template()
            input("\n계속하려면 Enter를 누르세요...")
        elif choice == 5:
            excel_input_mode()
        elif choice == 6:
            ai_prediction_mode()
        elif choice == 7:
            print("\n" + "="*50)
            print("미디어믹스 시뮬레이터를 종료합니다.")
            print("="*50)
            print("\n감사합니다! 👋")
            break


if __name__ == "__main__":
    main()

