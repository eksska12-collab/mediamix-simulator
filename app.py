"""
미디어믹스 시뮬레이터 - Streamlit 웹앱
Media Mix Simulator - Web Application

✅ 최종 테스트 체크리스트:
[ ] 홈 → 자동 생성 → 결과 확인
[ ] 홈 → 수동 입력 → 결과 확인
[ ] 홈 → 엑셀 업로드 → 결과 확인
[ ] 템플릿 다운로드 → 편집 → 업로드
[ ] 프리셋 저장 → 불러오기
[ ] 빠른 예산 조정 버튼
[ ] 목표 역산 계산
[ ] Excel 다운로드
[ ] 차트 표시
[ ] 인사이트 생성
[ ] 검증/추천 시스템
[ ] 에러 핸들링
"""

import io
import json
import os
import time
from datetime import datetime

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# 기존 모듈에서 필요한 함수 import
from media_mix_simulator import (
    apply_budget_adjustment,
    calculate_budget_competition_factor,
    calculate_media_performance,
    calculate_performance,
    create_scenario_dataframe,
    format_number,
    generate_scenarios,
    get_media_adjusted_metrics,
)

# 새로운 모듈에서 import
from modules import (
    # constants
    BENCHMARKS, INDUSTRY_BASE_METRICS, INDUSTRY_SEASON_WEIGHT,
    SEASONALITY_COMMON, SEASONALITY, MEDIA_MULTIPLIERS,
    MEDIA_CATEGORIES, ALL_MEDIA,
    RISK_RATIO_THRESHOLD, EFFICIENCY_WARNING_THRESHOLD,
    load_benchmarks_json, load_media_categories_json,
    get_available_industries, get_media_benchmarks,
    # calculations
    calculate_seasonality,
    estimate_conversion_increase,
    calculate_efficiency_grade,
    # validators
    validate_input,
    validate_efficiency,
    EFFICIENCY_RANGES,
    # excel_handler
    create_excel_download,
    # ui_components
    render_page_header,
    render_footer,
)

# =============================================================================
# 앱 설정
# =============================================================================

# 페이지 설정
st.set_page_config(
    page_title="미디어믹스 시뮬레이터 v1.0",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 프리셋 폴더 생성 (없으면 자동 생성)
if not os.path.exists('saved_presets'):
    os.makedirs('saved_presets')

# =============================================================================
# 추천 및 인사이트 함수
# Note: 이 함수들은 길어서 app.py에 유지합니다.
# 향후 리팩토링 시 modules/insights.py로 이동 가능
# =============================================================================

def generate_recommendations(scenarios, budget):
    """
    시뮬레이션 결과 기반 스마트 추천 생성 (구체적 수치 포함)
    
    Args:
        scenarios: 시나리오 데이터 (base 시나리오 사용)
        budget: 총 예산
    
    Returns:
        recommendations: 추천 리스트
    """
    recommendations = []
    
    # base 시나리오 데이터 사용
    media_data = scenarios.get('base', [])
    if not media_data:
        return recommendations
    
    # CPA 순으로 정렬
    sorted_media = sorted([m for m in media_data if m.get('cpa', 0) > 0], key=lambda x: x.get('cpa', 0))
    
    if len(sorted_media) >= 2:
        # 가장 효율 좋은 매체
        best_media = sorted_media[0]
        # 가장 효율 나쁜 매체
        worst_media = sorted_media[-1]
        
        # 핵심 추천: 비효율 매체 → 효율 매체로 예산 이동
        if worst_media['cpa'] > best_media['cpa'] * 1.5:
            # 비효율 매체의 10%를 효율 매체로 이동
            shift_ratio = worst_media.get('budget_ratio', 0) * 0.1
            shift_budget = budget * (shift_ratio / 100)
            
            # 효율 좋은 매체의 전환 데이터 기준 추가 전환 계산
            best_cvr = best_media.get('cvr', 0) / 100
            best_cpc = best_media.get('cpc', 0)
            if best_cpc > 0 and best_cvr > 0:
                additional_clicks = shift_budget / best_cpc
                additional_conversions = additional_clicks * best_cvr
                
                # CPA 개선 계산
                current_total_cv = sum(m.get('estimated_conversions_adjusted', 0) for m in media_data)
                current_avg_cpa = budget / current_total_cv if current_total_cv > 0 else 0
                new_total_cv = current_total_cv + additional_conversions
                new_avg_cpa = budget / new_total_cv if new_total_cv > 0 else 0
                cpa_improvement = current_avg_cpa - new_avg_cpa
                
                recommendations.append({
                    'type': 'info',
                    'icon': '💡',
                    'message': f"**{worst_media['name']}** 비중 {worst_media['budget_ratio']:.1f}% → **{worst_media['budget_ratio']-shift_ratio:.1f}%** 감소, "
                              f"**{best_media['name']}** 비중 {best_media['budget_ratio']:.1f}% → **{best_media['budget_ratio']+shift_ratio:.1f}%** 증가 시\n"
                              f"📈 전환 **+{additional_conversions:.0f}건**, 평균 CPA **-{cpa_improvement:,.0f}원** 개선 예상"
                })
    
    # 2. 비중이 낮지만 효율 좋은 매체 확대 제안
    for media in sorted_media[:2]:  # 상위 2개
        ratio = media.get('budget_ratio', 0)
        cpa = media.get('cpa', 0)
        name = media.get('name', '매체')
        cvr = media.get('cvr', 0) / 100
        cpc = media.get('cpc', 0)
        
        if ratio < 20 and cpa > 0:  # 비중이 20% 미만
            # 10%p 증가 시 효과
            increase_ratio = 10
            increase_budget = budget * (increase_ratio / 100)
            
            if cpc > 0 and cvr > 0:
                add_clicks = increase_budget / cpc
                add_cv = add_clicks * cvr
                add_cpa_impact = increase_budget / add_cv if add_cv > 0 else 0
                
                recommendations.append({
                    'type': 'info',
                    'icon': '🚀',
                    'message': f"**{name}** (현재 CPA {cpa:,.0f}원) 비중 {ratio:.1f}% → **{ratio+increase_ratio:.1f}%** 증가 시\n"
                              f"📈 전환 **+{add_cv:.0f}건**, 예상 CPA **{add_cpa_impact:,.0f}원** 유지"
                })
    
    # 3. 고비중 리스크 경고
    for media in media_data:
        ratio = media.get('budget_ratio', 0)
        name = media.get('name', '매체')
        
        if ratio > RISK_RATIO_THRESHOLD:
            recommendations.append({
                'type': 'warning',
                'icon': '⚠️',
                'message': f"**{name}** 의존도({ratio:.1f}%)가 높습니다. 매체 알고리즘 변경이나 정책 변화 시 리스크가 큽니다. "
                          f"다른 매체로 분산을 권장합니다."
            })
    
    # 4. ROAS 개선 기회
    low_roas_media = [m for m in media_data if 0 < m.get('roas', 0) < 150]
    if low_roas_media:
        for media in low_roas_media:
            name = media.get('name', '매체')
            roas = media.get('roas', 0)
            cpa = media.get('cpa', 0)
            revenue_per_cv = media.get('revenue_per_conversion', 0)
            
            # 목표 ROAS 150% 달성을 위한 필요 매출
            needed_revenue = cpa * 1.5
            current_revenue = revenue_per_cv
            revenue_gap = needed_revenue - current_revenue
            
            recommendations.append({
                'type': 'warning',
                'icon': '💰',
                'message': f"**{name}** ROAS {roas:.1f}%로 낮습니다. "
                          f"전환당 매출을 현재 {current_revenue:,.0f}원에서 **{needed_revenue:,.0f}원**으로 "
                          f"**(+{revenue_gap:,.0f}원)** 개선 시 ROAS 150% 달성 가능"
            })
    
    # 5. 전체 볼륨 확대 제안
    total_conversions = sum(m.get('estimated_conversions_adjusted', 0) for m in media_data)
    if 0 < total_conversions < 200:
        # 예산 30% 증액 시 효과
        new_budget = budget * 1.3
        estimated_new_cv = total_conversions * 1.3
        
        recommendations.append({
            'type': 'info',
            'icon': '📈',
            'message': f"현재 예상 전환수({total_conversions:.0f}건)가 부족합니다. "
                      f"총 예산을 {format_number(int(budget))}원 → **{format_number(int(new_budget))}원** "
                      f"**(+30%, +{format_number(int(new_budget - budget))}원)** 증액 시 "
                      f"전환 **{estimated_new_cv:.0f}건** 달성 예상"
        })
    
    return recommendations

def generate_ai_insights(result_data, industry, month, goal):
    """
    시뮬레이션 결과 기반 고급 AI 인사이트 생성
    
    Args:
        result_data: 시뮬레이션 결과 데이터
        industry: 업종
        month: 운영 월
        goal: 캠페인 목표
    
    Returns:
        insights: 인사이트 리스트
    """
    insights = []
    
    # 결과 데이터 구조에 따라 접근
    scenarios = result_data.get('scenarios', {})
    media_data = scenarios.get('base', []) if scenarios else result_data.get('media_list', [])
    
    # 총 전환수 계산
    total_conversions = sum(m.get('estimated_conversions_adjusted', m.get('conversions', 0)) for m in media_data)
    
    # 평균 CPA 계산
    total_budget = result_data.get('budget', 0)
    avg_cpa = (total_budget / total_conversions) if total_conversions > 0 else 0
    
    # 평균 ROAS 계산
    total_revenue = sum(m.get('total_revenue_adjusted', m.get('revenue', 0)) for m in media_data)
    avg_roas = (total_revenue / total_budget * 100) if total_budget > 0 else 0
    
    # 1. 성과 수준 평가
    if total_conversions >= 1000:
        insights.append({
            'type': 'success',
            'title': '🎯 우수한 전환 성과',
            'message': f'예상 전환수({total_conversions:,.0f}건)가 매우 높습니다. 안정적인 캠페인 운영이 가능합니다.'
        })
    elif total_conversions < 100:
        insights.append({
            'type': 'warning',
            'title': '⚠️ 전환 볼륨 부족',
            'message': f'예상 전환수({total_conversions:,.0f}건)가 적어 통계적 유의성이 낮을 수 있습니다. 예산 증액 또는 목표 조정을 권장합니다.'
        })
    
    # 2. 업종별 벤치마크 비교
    industry_avg_cpa = {
        '보험': 60000,
        '금융': 50000,
        '패션': 35000,
        'IT/테크': 45000
    }.get(industry, 50000)
    
    if avg_cpa > 0:
        if avg_cpa < industry_avg_cpa * 0.8:
            insights.append({
                'type': 'success',
                'title': '💰 효율적인 CPA',
                'message': f'평균 CPA({avg_cpa:,.0f}원)가 {industry} 업종 평균({industry_avg_cpa:,}원)보다 {((industry_avg_cpa - avg_cpa) / industry_avg_cpa * 100):.0f}% 낮습니다.'
            })
        elif avg_cpa > industry_avg_cpa * 1.3:
            insights.append({
                'type': 'error',
                'title': '📈 높은 CPA',
                'message': f'평균 CPA({avg_cpa:,.0f}원)가 {industry} 업종 평균({industry_avg_cpa:,}원)보다 높습니다. 타겟팅 또는 크리에이티브 개선이 필요합니다.'
            })
    
    # 3. 매체 다각화 분석
    sa_ratio = sum(m.get('budget_ratio', 0) for m in media_data if '검색' in m.get('category', ''))
    da_ratio = sum(m.get('budget_ratio', 0) for m in media_data if '디스플레이' in m.get('category', ''))
    
    if abs(sa_ratio - da_ratio) > RISK_RATIO_THRESHOLD:
        dominant = "검색광고" if sa_ratio > da_ratio else "디스플레이광고"
        insights.append({
            'type': 'info',
            'title': '🎯 매체 편중',
            'message': f'{dominant} 비중이 높습니다. 균형잡힌 믹스를 위해 다른 매체 확대를 검토하세요.'
        })
    
    # 4. 계절성 활용
    season_factor = SEASONALITY_COMMON.get(month, 1.0)
    if season_factor >= 1.15:
        insights.append({
            'type': 'success',
            'title': '🔥 최적의 시기',
            'message': f'{month}월은 {industry} 업종의 성수기입니다. 공격적인 집행을 권장합니다.'
        })
    elif season_factor <= 0.85:
        insights.append({
            'type': 'warning',
            'title': '❄️ 비수기 대응',
            'message': f'{month}월은 효율이 낮은 시기입니다. 브랜딩 중심 또는 예산 축소를 고려하세요.'
        })
    
    # 5. 목표 일치도
    if goal and "전환" in str(goal) and sa_ratio < 60:
        insights.append({
            'type': 'info',
            'title': '🎯 목표-믹스 불일치',
            'message': '전환 중심 목표인데 검색광고 비중이 낮습니다. SA 비중을 60% 이상으로 증가시키면 더 좋은 성과를 기대할 수 있습니다.'
        })
    
    # 6. ROAS 평가
    if avg_roas > 0:
        if avg_roas >= 200:
            insights.append({
                'type': 'success',
                'title': '💰 높은 수익성',
                'message': f'평균 ROAS({avg_roas:.1f}%)가 우수합니다. 예산 증액을 고려해보세요.'
            })
        elif avg_roas < 100:
            insights.append({
                'type': 'error',
                'title': '⚠️ 낮은 수익성',
                'message': f'평균 ROAS({avg_roas:.1f}%)가 100% 미만입니다. 전환당 매출 증가 또는 CPA 개선이 필요합니다.'
            })
    
    return insights

# 세션 스테이트 초기화
if 'results' not in st.session_state:
    st.session_state.results = None
if 'budget' not in st.session_state:
    st.session_state.budget = 10000000
if 'selected_mode' not in st.session_state:
    st.session_state.selected_mode = "🏠 홈"
if 'comparison_mode' not in st.session_state:
    st.session_state.comparison_mode = False
if 'previous_result' not in st.session_state:
    st.session_state.previous_result = None
if 'media_ratios' not in st.session_state:
    st.session_state.media_ratios = []

# =============================================================================
# CSS 스타일링 통합
# =============================================================================
st.markdown("""
<style>
/* 카드 스타일 개선 */
.stAlert {
    border-radius: 8px;
}

/* 메트릭 카드 개선 */
div[data-testid="stMetricValue"] {
    font-size: 24px;
    font-weight: 600;
}
</style>
""", unsafe_allow_html=True)

# 사이드바 메뉴
st.sidebar.title("🎯 모드 선택")
mode_list = ["🏠 홈", "📊 자동 생성", "✏️ 수동 입력", "🎯 목표 역산 계산", "📁 엑셀 업로드", "📥 템플릿 다운로드"]

# 기본 인덱스 설정 (session_state에서 현재 모드 찾기)
if st.session_state.selected_mode in mode_list:
    default_index = mode_list.index(st.session_state.selected_mode)
else:
    default_index = 0
    st.session_state.selected_mode = mode_list[0]

# radio 버튼 (session_state와 동기화)
mode = st.sidebar.radio(
    "작업 모드를 선택하세요:",
    mode_list,
    index=default_index
)

# radio 선택이 변경되었을 때만 session_state 업데이트
if mode != st.session_state.selected_mode:
    st.session_state.selected_mode = mode
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.info("💡 **Tip**: 자동 생성 모드로 빠르게 시작하세요!")

# 사이드바 프리셋 관리 섹션
st.sidebar.markdown("---")
st.sidebar.markdown("### 📁 프리셋 관리")

# 저장된 프리셋 목록
preset_files = []
if os.path.exists('saved_presets'):
    preset_files = [f.replace('.json', '') for f in os.listdir('saved_presets') if f.endswith('.json')]

if preset_files:
    selected_preset = st.sidebar.selectbox(
        "저장된 프리셋",
        ["선택하세요..."] + preset_files,
        key="preset_selector"
    )
    
    # 선택된 프리셋 미리보기
    if selected_preset != "선택하세요...":
        try:
            with open(f'saved_presets/{selected_preset}.json', 'r', encoding='utf-8') as f:
                preview = json.load(f)
            
            st.sidebar.caption(f"📊 모드: {preview.get('mode', 'N/A')}")
            st.sidebar.caption(f"💰 예산: {preview.get('budget', 0):,}원")
            if 'industry' in preview:
                st.sidebar.caption(f"🏢 업종: {preview['industry']}")
            st.sidebar.caption(f"📅 저장: {preview.get('saved_at', 'N/A')}")
        except Exception as e:
            st.sidebar.error(f"미리보기 로드 실패: {e}")
    
    col1, col2 = st.sidebar.columns(2)
    with col1:
        if st.button("불러오기", key="load_preset", use_container_width=True):
            if selected_preset != "선택하세요...":
                try:
                    with open(f'saved_presets/{selected_preset}.json', 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    # 세션 상태에 저장
                    if data.get('mode') == 'AI':
                        st.session_state.preset_budget = data.get('budget', 50000000)
                        st.session_state.preset_industry = data.get('industry', '보험')
                        st.session_state.preset_month = data.get('month', 1)
                        st.session_state.preset_goal = data.get('goal', '균형')
                        st.session_state.preset_media_config = data.get('media_config', {})
                        st.session_state.selected_mode = "📊 자동 생성"
                    
                    elif data.get('mode') == 'Manual':
                        st.session_state.preset_budget = data.get('budget', 50000000)
                        st.session_state.preset_media_data = data.get('media_data', [])
                        st.session_state.selected_mode = "✏️ 수동 입력"
                    
                    st.success(f"✅ '{selected_preset}' 프리셋 불러오기 완료!")
                    st.rerun()
                except Exception as e:
                    st.sidebar.error(f"불러오기 실패: {e}")
            else:
                st.sidebar.warning("프리셋을 선택하세요")
    
    with col2:
        if st.button("삭제", key="delete_preset", use_container_width=True):
            if selected_preset != "선택하세요...":
                try:
                    os.remove(f'saved_presets/{selected_preset}.json')
                    st.success(f"✅ '{selected_preset}' 프리셋 삭제 완료!")
                    st.rerun()
                except Exception as e:
                    st.sidebar.error(f"삭제 실패: {e}")
            else:
                st.sidebar.warning("프리셋을 선택하세요")

else:
    st.sidebar.info("💡 저장된 프리셋이 없습니다")

# 사이드바 브랜딩 및 정보
st.sidebar.markdown("---")
with st.sidebar.expander("⌨️ 단축키"):
    st.markdown("""
    - `Ctrl + Enter`: 실행
    - `Ctrl + R`: 새로고침
    - `Esc`: 팝업 닫기
    """)

st.sidebar.markdown("""
---
<div style="text-align: center; padding: 20px;">
    <p style="color: #757575; font-size: 12px; margin: 0;">Powered by</p>
    <h3 style="color: #2196F3; margin: 5px 0;">by JH</h3>
    <p style="color: #757575; font-size: 11px;">Performance Marketing 10 Team</p>
</div>
""", unsafe_allow_html=True)

# 버전 정보 (실시간 날짜 반영)
current_date = datetime.now().strftime("%Y.%m.%d")
st.sidebar.caption(f"v2.5.0 | {current_date}")


# ===== 홈 화면 =====
if mode == "🏠 홈":
    # 공통 헤더
    render_page_header()
    
    # 환영 메시지
    st.markdown("""
    ### 👋 환영합니다!
    JH Performance Marketing 10 Team 전용 미디어믹스 예산 배분 및 성과 예측 툴입니다.
    """)
    
    st.markdown("### 🚀 빠른 시작")
    
    # 3개 카드 레이아웃
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 10px;
                    border: 1px solid #e0e0e0; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
                    margin-bottom: 10px;">
            <h3 style="color: #2196F3; margin-top: 0;">📊 자동 생성</h3>
            <p style="color: #1a1a1a;"><strong>추천 대상:</strong> 빠른 초안 작성</p>
            <ul style="color: #424242; margin: 10px 0;">
                <li>⏱️ 소요시간: 30초</li>
                <li>🎯 정확도: ⭐⭐⭐⭐</li>
                <li>💡 업종별 벤치마크 데이터로 빠른 초안 작성</li>
            </ul>
            <p style="color: #1a1a1a;"><strong>이런 때 사용:</strong><br/>
            <span style="color: #424242;">
            - 클라이언트 제안 초안<br/>
            - 빠른 예산 시뮬레이션<br/>
            - 업종 평균 기준 예측
            </span>
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("📊 자동 생성 시작", key="btn_ai", use_container_width=True):
            st.session_state.selected_mode = "📊 자동 생성"
            st.rerun()
    
    with col2:
        st.markdown("""
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 10px;
                    border: 1px solid #e0e0e0; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
                    margin-bottom: 10px;">
            <h3 style="color: #2196F3; margin-top: 0;">✏️ 수동 입력</h3>
            <p style="color: #1a1a1a;"><strong>추천 대상:</strong> 정밀한 예측</p>
            <ul style="color: #424242; margin: 10px 0;">
                <li>⏱️ 소요시간: 3분</li>
                <li>🎯 정확도: ⭐⭐⭐⭐⭐</li>
                <li>💡 실제 효율 데이터 반영</li>
            </ul>
            <p style="color: #1a1a1a;"><strong>이런 때 사용:</strong><br/>
            <span style="color: #424242;">
            - 실제 캠페인 데이터 보유<br/>
            - 정밀한 성과 예측 필요<br/>
            - 매체별 상세 분석
            </span>
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("✏️ 수동 입력 시작", key="btn_manual", use_container_width=True):
            st.session_state.selected_mode = "✏️ 수동 입력"
            st.rerun()
    
    with col3:
        st.markdown("""
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 10px;
                    border: 1px solid #e0e0e0; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
                    margin-bottom: 10px;">
            <h3 style="color: #2196F3; margin-top: 0;">📁 엑셀 업로드</h3>
            <p style="color: #1a1a1a;"><strong>추천 대상:</strong> 대량 처리</p>
            <ul style="color: #424242; margin: 10px 0;">
                <li>⏱️ 소요시간: 1분</li>
                <li>🎯 정확도: ⭐⭐⭐⭐⭐</li>
                <li>💡 템플릿 기반 입력</li>
            </ul>
            <p style="color: #1a1a1a;"><strong>이런 때 사용:</strong><br/>
            <span style="color: #424242;">
            - 반복적인 시뮬레이션<br/>
            - 여러 클라이언트 동시 작업<br/>
            - 데이터 재사용
            </span>
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("📁 엑셀 업로드", key="btn_excel", use_container_width=True):
            st.session_state.selected_mode = "📁 엑셀 업로드"
            st.rerun()
    
    st.markdown("---")
    
    # 주요 기능 안내
    st.markdown("### ✨ 주요 기능")
    
    feat_col1, feat_col2 = st.columns(2)
    
    with feat_col1:
        st.markdown("""
        - ✅ **업종별 자동 생성**: 보험/패션/금융/IT 벤치마크
        - ✅ **3개 시나리오 비교**: 보수/기본/공격
        - ✅ **실시간 차트**: 파이차트/바차트
        - ✅ **Excel 다운로드**: 즉시 결과 저장
        """)
    
    with feat_col2:
        st.markdown("""
        - ✅ **프리셋 관리**: 자주 쓰는 설정 저장
        - ✅ **스마트 검증**: 비현실적 효율 경고
        - ✅ **계절성 보정**: 월별 자동 조정
        - ✅ **빠른 조정**: ±10%, ±30% 원클릭
        """)
    
    st.markdown("---")
    
    # 버전 정보 및 업데이트 내역
    with st.expander("📋 버전 정보 및 업데이트 내역"):
        current_version_date = datetime.now().strftime("%Y.%m.%d")
        st.markdown(f"""
        **v2.5** ({current_version_date})
        - 🆕 라이트모드 UI 적용 (흰색 배경, 파란색 강조)
        - 🆕 핵심 지표 카드 4개 추가 (전환수/CPA/ROAS/효율등급)
        - 🆕 매체 효율 순위 시스템 (Best 3 / Worst 3)
        - 🆕 구체적 개선 제안 (수치 기반 추천)
        - 🆕 팀 공유 모드 (run_webapp_share.bat)
        - 🆕 효율 등급 시스템 (S/A/B/C)
        - 🎨 홈 카드 높이 자동 조정
        
        **v1.0** (2025.12.30)
        - 최초 릴리즈
        - 업종/계절성 보정 시스템
        - INDUSTRY_BASE_METRICS 기반 정밀 예측
        - 예산 규모별 경쟁도 보정
        - 매체별 Multiplier 시스템
        - 홈 대시보드
        - 계절성 보정 기능
        - Excel 템플릿 다운로드 기능
        - 시나리오 비교 차트
        - 자동 생성 기능
        - 수동 입력 모드
        - Excel 업로드 기능
        - 매체 효율 순위 시스템 (Best 3 / Worst 3)
        - 구체적 개선 제안 (수치 기반 추천)
        - 팀 공유 모드 (run_webapp_share.bat)
        - 효율 등급 시스템 (S/A/B/C)
        """)


# ===== 자동 생성 모드 =====
elif mode == "📊 자동 생성":
    # 공통 헤더
    render_page_header("자동 생성")
    st.markdown("업종별 벤치마크 데이터를 기반으로 최적의 미디어믹스를 자동 생성합니다.")
    
    # UI 버전 관리 (위젯 재생성용)
    if 'ui_rev' not in st.session_state:
        st.session_state.ui_rev = 0
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📝 기본 정보")
        
        # 프리셋에서 불러온 값이 있으면 적용
        default_budget = st.session_state.get('preset_budget', 50000000)
        
        budget = st.number_input(
            "💰 총 예산 (원)",
            min_value=1000000,
            max_value=2000000000,
            value=default_budget,
            step=1000000,
            format="%d",
            key="ai_budget_input",
            help="""
            📌 Tips:
            • 1천만원 이상 권장
            • 예산이 클수록 CPC 경쟁도 상승
            • 최대 20억원까지 입력 가능
            """
        )
        st.session_state.budget = budget
        
        # 업종 선택 (프리셋 적용)
        industries_list = get_available_industries()
        default_industry = st.session_state.get('preset_industry', industries_list[0] if industries_list else '보험')
        industry_index = industries_list.index(default_industry) if default_industry in industries_list else 0
        industry = st.selectbox(
            "🏢 업종 선택",
            industries_list,
            index=industry_index,
            help="""
            📌 업종별 특징:
            • 보험: 높은 CPC, 성수기 11-12월
            • 금융: 안정적 효율, 연초 성수기
            • 패션: 계절별 변동 큼, 시즌 세일 중요
            • IT/테크: 계절성 적음, 안정적 운영
            """
        )
        
        # 월 선택 (프리셋 적용)
        default_month = st.session_state.get('preset_month', 1)
        month_index = default_month - 1 if 1 <= default_month <= 12 else 0
        month = st.selectbox(
            "운영 월",
            list(range(1, 13)),
            index=month_index,
            format_func=lambda x: f"{x}월 (계절성 보정: {SEASONALITY[x]:.2f}x)"
        )
        
        # 목표 선택 (프리셋 적용)
        goals_list = ["인지도 (DA 중심)", "전환 (SA 중심)", "균형 (SA/DA 균형)"]
        default_goal = st.session_state.get('preset_goal', '균형')
        # 목표 문자열 매칭
        goal_map = {
            "인지도 (DA 중심)": "인지도",
            "전환 (SA 중심)": "전환",
            "균형 (SA/DA 균형)": "균형"
        }
        goal_index = 2  # 기본값: 균형
        for idx, goal_option in enumerate(goals_list):
            if goal_map[goal_option] in default_goal or default_goal in goal_option:
                goal_index = idx
                break
        
        goal = st.radio(
            "마케팅 목표",
            goals_list,
            index=goal_index
        )
    
    with col2:
        st.subheader("📊 매체 선택 및 예산 배분")
        
        # 매체 목록 매핑 (당근은 SA와 DA가 별도 키 필요)
        media_display_map = {
            '네이버 검색광고': '네이버_SA',
            '구글 검색광고': '구글_SA',
            '카카오 검색광고': '카카오_SA',
            '네이버 파워컨텐츠': '네이버파워컨텐츠_SA',
            '메타': '메타_DA',
            '구글 배너광고': '구글GDA_DA',
            '틱톡': '틱톡_DA',
            '네이버 GFA': '네이버GFA_DA',
            '카카오 모먼트': '카카오모먼트_DA',
            '크리테오': '크리테오_DA',
            '토스': '토스_DA',
            '카카오페이': '카카오페이_DA'
        }
        
        # 당근은 카테고리에 따라 동적으로 처리
        def get_media_key(media_name, category):
            if media_name == '당근':
                return '당근_SA' if category == 'SA' else '당근_DA'
            return media_display_map.get(media_name)
        
        # 모든 매체 표시 (BENCHMARKS에 없으면 시뮬레이션 시 안내)
        benchmark_media = list(BENCHMARKS.get(industry, {}).keys())
        available_sa = MEDIA_CATEGORIES['SA']
        available_da = MEDIA_CATEGORIES['DA']
        
        # session_state 초기화 (프리셋에서 불러온 경우 적용)
        if 'media_config' not in st.session_state:
            st.session_state.media_config = {}
        
        # 프리셋에서 불러온 media_config 적용
        if 'preset_media_config' in st.session_state and st.session_state.preset_media_config:
            st.session_state.media_config = st.session_state.preset_media_config.copy()
            del st.session_state.preset_media_config
        
        # 자동 균등 배분 버튼
        col_auto1, col_auto2 = st.columns([1, 1])
        with col_auto1:
            if st.button("⚡ 균등 배분", key="equal_distribute_btn", use_container_width=True, help="선택된 매체들을 균등하게 배분합니다"):
                # 활성화된 매체들 찾기
                active_media = [m for m in list(available_sa) + list(available_da) 
                               if st.session_state.media_config.get(f"{m}_active", True)]
                
                if active_media:
                    equal_ratio = 100.0 / len(active_media)
                    for media in active_media:
                        # media_config만 업데이트 (source of truth)
                        st.session_state.media_config[f"{media}_ratio"] = equal_ratio
                else:
                    # 모든 매체를 활성화하고 균등 배분
                    all_media = list(available_sa) + list(available_da)
                    equal_ratio = 100.0 / len(all_media)
                    for media in all_media:
                        # media_config만 업데이트 (source of truth)
                        st.session_state.media_config[f"{media}_active"] = True
                        st.session_state.media_config[f"{media}_ratio"] = equal_ratio
                
                # UI 재생성을 위해 버전 증가
                st.session_state.ui_rev += 1
                st.rerun()
        
        with col_auto2:
            if st.button("🔄 초기화", key="reset_config_btn", use_container_width=True, help="모든 설정을 초기화합니다"):
                # 모든 매체 목록
                all_media = list(available_sa) + list(available_da)
                default_ratio = 100.0 / len(all_media)
                
                # media_config만 초기화 (source of truth)
                st.session_state.media_config = {}
                for media in all_media:
                    st.session_state.media_config[f"{media}_active"] = True
                    st.session_state.media_config[f"{media}_ratio"] = default_ratio
                
                # UI 재생성을 위해 버전 증가
                st.session_state.ui_rev += 1
                st.rerun()
        
        st.markdown("---")
        
        # SA 섹션
        with st.expander("🔍 검색광고 (SA)", expanded=True):
            for media in available_sa:
                col1, col2, col3 = st.columns([1.5, 2.5, 1])
                
                # 기본값: 모든 매체 활성화
                default_active = st.session_state.media_config.get(f"{media}_active", True)
                default_ratio = st.session_state.media_config.get(f"{media}_ratio", 100.0 / (len(available_sa) + len(available_da)))
                
                # 벤치마크 데이터 확인
                media_key = get_media_key(media, 'SA')
                has_benchmark = media_key and media_key in benchmark_media
                
                with col1:
                    # ui_rev 포함으로 위젯 재생성
                    ui_rev = st.session_state.ui_rev
                    active = st.toggle(
                        media,
                        value=default_active,
                        key=f"ai_sa_{media}_toggle_{ui_rev}"
                    )
                    st.session_state.media_config[f"{media}_active"] = active
                    
                    # 벤치마크 데이터 없으면 표시
                    if not has_benchmark:
                        st.caption("⚠️ 데이터 없음")
                
                if active:
                    # 키 정의 (ui_rev 포함으로 위젯 재생성)
                    ui_rev = st.session_state.ui_rev
                    slider_key = f"ai_sa_{media}_slider_{ui_rev}"
                    input_key = f"ai_sa_{media}_input_{ui_rev}"
                    config_key = f"{media}_ratio"
                    
                    # media_config에서 현재 비율 가져오기 (source of truth, 매체별)
                    current_ratio = st.session_state.media_config.get(config_key, default_ratio)
                    
                    # 범위 강제 (0~100)
                    current_ratio = max(0.0, min(100.0, current_ratio))
                    
                    # 콜백: 값 변경 시 해당 매체의 media_config만 업데이트 (ui_rev 증가 없음)
                    def make_slider_callback(s_key, c_key):
                        def callback():
                            val = st.session_state.get(s_key)
                            if val is not None:
                                # 범위 강제 (0~100)
                                val = max(0.0, min(100.0, val))
                                # 해당 매체의 media_config만 업데이트
                                st.session_state.media_config[c_key] = val
                        return callback
                    
                    def make_input_callback(i_key, c_key):
                        def callback():
                            val = st.session_state.get(i_key)
                            if val is not None:
                                # 범위 강제 (0~100)
                                val = max(0.0, min(100.0, val))
                                # 해당 매체의 media_config만 업데이트
                                st.session_state.media_config[c_key] = val
                        return callback
                    
                    with col2:
                        st.slider(
                            "비중 (%)",
                            min_value=0.0,
                            max_value=100.0,
                            value=current_ratio,
                            step=1.0,
                            key=slider_key,
                            label_visibility="collapsed",
                            on_change=make_slider_callback(slider_key, config_key)
                        )
                    
                    with col3:
                        st.number_input(
                            "비중",
                            min_value=0.0,
                            max_value=100.0,
                            value=current_ratio,
                            step=1.0,
                            key=input_key,
                            label_visibility="collapsed",
                            on_change=make_input_callback(input_key, config_key)
                        )
                    
                    # media_config에 최종 값 동기화 (source of truth 유지, 매체별)
                    if slider_key in st.session_state:
                        st.session_state.media_config[config_key] = st.session_state[slider_key]
                    elif input_key in st.session_state:
                        st.session_state.media_config[config_key] = st.session_state[input_key]
                else:
                    with col2:
                        st.markdown("*<span style='color: gray;'>비활성화됨</span>*", unsafe_allow_html=True)
                    with col3:
                        st.markdown("")
                    st.session_state.media_config[f"{media}_ratio"] = 0.0
        
        # DA 섹션
        with st.expander("📱 디스플레이 (DA)", expanded=True):
            for media in available_da:
                col1, col2, col3 = st.columns([1.5, 2.5, 1])
                
                # 기본값: 모든 매체 활성화
                default_active = st.session_state.media_config.get(f"{media}_active", True)
                default_ratio = st.session_state.media_config.get(f"{media}_ratio", 100.0 / (len(available_sa) + len(available_da)))
                
                # 벤치마크 데이터 확인
                media_key = get_media_key(media, 'DA')
                has_benchmark = media_key and media_key in benchmark_media
                
                with col1:
                    # ui_rev 포함으로 위젯 재생성
                    ui_rev = st.session_state.ui_rev
                    active = st.toggle(
                        media,
                        value=default_active,
                        key=f"ai_da_{media}_toggle_{ui_rev}"
                    )
                    st.session_state.media_config[f"{media}_active"] = active
                    
                    # 벤치마크 데이터 없으면 표시
                    if not has_benchmark:
                        st.caption("⚠️ 데이터 없음")
                
                if active:
                    # 키 정의 (ui_rev 포함으로 위젯 재생성)
                    ui_rev = st.session_state.ui_rev
                    slider_key = f"ai_da_{media}_slider_{ui_rev}"
                    input_key = f"ai_da_{media}_input_{ui_rev}"
                    config_key = f"{media}_ratio"
                    
                    # media_config에서 현재 비율 가져오기 (source of truth, 매체별)
                    current_ratio = st.session_state.media_config.get(config_key, default_ratio)
                    
                    # 범위 강제 (0~100)
                    current_ratio = max(0.0, min(100.0, current_ratio))
                    
                    # 콜백: 값 변경 시 해당 매체의 media_config만 업데이트 (ui_rev 증가 없음)
                    def make_slider_callback(s_key, c_key):
                        def callback():
                            val = st.session_state.get(s_key)
                            if val is not None:
                                # 범위 강제 (0~100)
                                val = max(0.0, min(100.0, val))
                                # 해당 매체의 media_config만 업데이트
                                st.session_state.media_config[c_key] = val
                        return callback
                    
                    def make_input_callback(i_key, c_key):
                        def callback():
                            val = st.session_state.get(i_key)
                            if val is not None:
                                # 범위 강제 (0~100)
                                val = max(0.0, min(100.0, val))
                                # 해당 매체의 media_config만 업데이트
                                st.session_state.media_config[c_key] = val
                        return callback
                    
                    with col2:
                        st.slider(
                            "비중 (%)",
                            min_value=0.0,
                            max_value=100.0,
                            value=current_ratio,
                            step=1.0,
                            key=slider_key,
                            label_visibility="collapsed",
                            on_change=make_slider_callback(slider_key, config_key)
                        )
                    
                    with col3:
                        st.number_input(
                            "비중",
                            min_value=0.0,
                            max_value=100.0,
                            value=current_ratio,
                            step=1.0,
                            key=input_key,
                            label_visibility="collapsed",
                            on_change=make_input_callback(input_key, config_key)
                        )
                    
                    # media_config에 최종 값 동기화 (source of truth 유지, 매체별)
                    if slider_key in st.session_state:
                        st.session_state.media_config[config_key] = st.session_state[slider_key]
                    elif input_key in st.session_state:
                        st.session_state.media_config[config_key] = st.session_state[input_key]
                else:
                    with col2:
                        st.markdown("*<span style='color: gray;'>비활성화됨</span>*", unsafe_allow_html=True)
                    with col3:
                        st.markdown("")
                    st.session_state.media_config[f"{media}_ratio"] = 0.0
        
        st.markdown("---")
        
        # 실시간 합계 표시
        total_ratio = sum(
            st.session_state.media_config.get(f"{m}_ratio", 0.0)
            for m in list(available_sa) + list(available_da)
            if st.session_state.media_config.get(f"{m}_active", True)
        )
        
        active_count = sum(
            1 for m in list(available_sa) + list(available_da)
            if st.session_state.media_config.get(f"{m}_active", True)
        )
        
        # 허용 오차 (0.1%)
        TOLERANCE = 0.1
        is_valid_ratio = abs(total_ratio - 100.0) < TOLERANCE
        
        # 합계 표시
        if is_valid_ratio:
            st.success(f"✅ 총 비중: **{total_ratio:.1f}%** (선택: {active_count}개 매체)")
        elif total_ratio > 100.0:
            st.error(f"⚠️ 총 비중: **{total_ratio:.1f}%** - 100%를 초과했습니다! (초과: {total_ratio - 100:.1f}%)")
        elif total_ratio < 100.0:
            st.warning(f"⚠️ 총 비중: **{total_ratio:.1f}%** - 100%에 맞춰주세요! (부족: {100 - total_ratio:.1f}%)")
        else:
            st.info(f"📊 총 비중: **{total_ratio:.1f}%**")
        
        # 자동 정규화 옵션
        st.markdown("---")
        normalize_col1, normalize_col2 = st.columns([2, 1])
        
        with normalize_col1:
            auto_normalize = st.toggle(
                "🔧 자동 정규화 (비중 합계를 자동으로 100%로 맞춤)",
                value=st.session_state.get('auto_normalize', True),
                key="auto_normalize_toggle",
                help="ON: 실행 시 비중 합계를 자동으로 100%로 조정합니다"
            )
            st.session_state['auto_normalize'] = auto_normalize
        
        with normalize_col2:
            if not is_valid_ratio and st.button("⚡ 지금 정규화", use_container_width=True, help="현재 비중을 100%로 즉시 조정"):
                # 활성 매체 목록
                active_media_list = [
                    m for m in list(available_sa) + list(available_da)
                    if st.session_state.media_config.get(f"{m}_active", True)
                ]
                
                if active_media_list:
                    if total_ratio == 0:
                        # 합계가 0이면 균등 배분
                        equal_ratio = 100.0 / len(active_media_list)
                        for media in active_media_list:
                            # media_config만 업데이트 (source of truth)
                            st.session_state.media_config[f"{media}_ratio"] = equal_ratio
                    else:
                        # 비례 정규화 후 반올림 보정
                        normalized_ratios = []
                        for media in active_media_list:
                            current = st.session_state.media_config.get(f"{media}_ratio", 0.0)
                            normalized = (current / total_ratio) * 100.0
                            rounded = round(normalized, 1)
                            normalized_ratios.append((media, rounded))
                        
                        # 반올림 후 합계 확인 및 보정
                        current_sum = sum(r for _, r in normalized_ratios)
                        if abs(current_sum - 100.0) > 0.01:
                            # 마지막 항목에 잔여분 보정
                            diff = 100.0 - current_sum
                            normalized_ratios[-1] = (normalized_ratios[-1][0], normalized_ratios[-1][1] + diff)
                        
                        # media_config만 업데이트 (source of truth)
                        for media, normalized_ratio in normalized_ratios:
                            st.session_state.media_config[f"{media}_ratio"] = normalized_ratio
                
                # UI 재생성을 위해 버전 증가
                st.session_state.ui_rev += 1
                st.rerun()
        
        st.markdown("---")
        st.info(f"📊 예산: **{format_number(budget)}원**")
        st.info(f"🏢 업종: **{industry}** / 📅 {month}월 (계절성: {SEASONALITY[month]:.2f}x)")
    
    # 실시간 예상 결과 미리보기
    if budget and industry and month:
        st.markdown("---")
        st.markdown("### 💡 예상 결과 미리보기")
        
        # 간단한 사전 계산
        base_metrics = INDUSTRY_BASE_METRICS[industry]
        season_factor = SEASONALITY_COMMON[month]
        
        # 업종별 계절성 가중치 적용
        industry_season = INDUSTRY_SEASON_WEIGHT[industry]
        if month in industry_season['high_months']:
            season_factor *= industry_season['high_multiplier']
        elif month in industry_season['low_months']:
            season_factor *= industry_season['low_multiplier']
        
        # 대략적인 성과 예측
        est_cpc = base_metrics['CPC'] * calculate_budget_competition_factor(budget)
        est_ctr = base_metrics['CTR'] * season_factor
        est_cvr = base_metrics['CVR'] * season_factor
        
        est_clicks = budget / est_cpc
        est_conversions = est_clicks * est_cvr
        est_cpa = budget / est_conversions if est_conversions > 0 else 0
        
        # 미리보기 카드
        preview_col1, preview_col2, preview_col3, preview_col4 = st.columns(4)
        
        with preview_col1:
            st.metric(
                "예상 전환수",
                f"~{est_conversions:,.0f}건",
                help="업종 평균 기준 대략적 예측"
            )
        
        with preview_col2:
            st.metric(
                "예상 CPA",
                f"~{est_cpa:,.0f}원"
            )
        
        with preview_col3:
            st.metric(
                "예상 클릭수",
                f"~{est_clicks:,.0f}회"
            )
        
        with preview_col4:
            # 계절성 표시
            season_status = "높음" if month in industry_season.get('high_months', []) else \
                           "낮음" if month in industry_season.get('low_months', []) else "보통"
            
            st.metric(
                "계절성",
                season_status,
                f"{(season_factor - 1) * 100:+.0f}%"
            )
        
        # 간단한 인사이트
        insights = []
        
        if season_status == "높음":
            insights.append("🔥 성수기입니다. 평소보다 높은 성과가 예상됩니다.")
        elif season_status == "낮음":
            insights.append("❄️ 비수기입니다. 효율이 다소 낮을 수 있습니다.")
        
        if budget >= 100_000_000:
            insights.append("💰 대규모 예산으로 CPC 경쟁도가 높아질 수 있습니다.")
        
        if est_conversions < 100:
            insights.append("⚠️ 예상 전환수가 적습니다. 예산 증액 검토를 권장합니다.")
        
        if insights:
            with st.expander("📊 즉시 인사이트"):
                for insight in insights:
                    st.caption(insight)
    
    # 버튼 활성화 조건 확인
    total_ratio_check = sum(
        st.session_state.media_config.get(f"{m}_ratio", 0.0)
        for m in list(available_sa) + list(available_da)
        if st.session_state.media_config.get(f"{m}_active", True)
    )
    active_media_check = [
        m for m in list(available_sa) + list(available_da)
        if st.session_state.media_config.get(f"{m}_active", True)
    ]
    
    TOLERANCE = 0.1
    can_execute = abs(total_ratio_check - 100.0) < TOLERANCE and len(active_media_check) > 0
    auto_normalize_enabled = st.session_state.get('auto_normalize', True)
    
    # 비활성화 조건 시 경고 표시
    if not can_execute and not auto_normalize_enabled:
        st.warning("⚠️ 자동 정규화가 OFF 상태이고 비중 합계가 100%가 아닙니다. 실행하려면 비중을 100%로 맞추거나 자동 정규화를 ON하세요.")
    
    if st.button("🚀 시뮬레이션 실행", type="primary", use_container_width=True, disabled=not can_execute and not auto_normalize_enabled):
        # 비중 합계 검증
        total_ratio = sum(
            st.session_state.media_config.get(f"{m}_ratio", 0.0)
            for m in list(available_sa) + list(available_da)
            if st.session_state.media_config.get(f"{m}_active", True)
        )
        
        active_media = [
            m for m in list(available_sa) + list(available_da)
            if st.session_state.media_config.get(f"{m}_active", True)
        ]
        
        if not active_media:
            st.error("⚠️ 최소 1개 이상의 매체를 선택해주세요!")
            st.stop()
        
        # 자동 정규화 적용 (필요시)
        if abs(total_ratio - 100.0) >= TOLERANCE:
            if auto_normalize_enabled:
                st.info("🔧 자동 정규화 적용 중...")
                
                if total_ratio == 0:
                    # 합계가 0이면 균등 배분
                    equal_ratio = 100.0 / len(active_media)
                    for media in active_media:
                        st.session_state.media_config[f"{media}_ratio"] = equal_ratio
                else:
                    # 비례 정규화 후 반올림 보정
                    normalized_ratios = []
                    for media in active_media:
                        current = st.session_state.media_config.get(f"{media}_ratio", 0.0)
                        normalized = (current / total_ratio) * 100.0
                        rounded = round(normalized, 1)
                        normalized_ratios.append((media, rounded))
                    
                    # 반올림 후 합계 확인 및 보정
                    current_sum = sum(r for _, r in normalized_ratios)
                    if abs(current_sum - 100.0) > 0.01:
                        # 마지막 항목에 잔여분 보정
                        diff = 100.0 - current_sum
                        normalized_ratios[-1] = (normalized_ratios[-1][0], normalized_ratios[-1][1] + diff)
                    
                    # 적용
                    for media, normalized_ratio in normalized_ratios:
                        st.session_state.media_config[f"{media}_ratio"] = normalized_ratio
                
                # 재계산
                total_ratio = 100.0
                st.success("✅ 비중이 100%로 자동 조정되었습니다!")
            else:
                st.error(f"⚠️ 매체 비중의 합이 {total_ratio:.1f}%입니다. 정확히 100%로 맞춰주세요!")
                st.stop()
        
        # 로딩 애니메이션 개선
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("📊 업종 데이터 로딩...")
            progress_bar.progress(20)
            time.sleep(0.3)
            
            status_text.text("🔧 계절성 보정 적용...")
            progress_bar.progress(40)
            time.sleep(0.3)
            
            status_text.text("💰 예산 배분 계산...")
            progress_bar.progress(60)
            time.sleep(0.3)
            
            status_text.text("📈 성과 예측 중...")
            progress_bar.progress(80)
            
            # 목표별 타입 추출 (인사이트 생성용)
            if "인지도" in goal:
                goal_name = "인지도"
            elif "전환" in goal:
                goal_name = "전환"
            else:
                goal_name = "균형"
            
            # 매체 선택 및 데이터 생성
            selected_media = []
            skipped_media = []  # BENCHMARKS에 없어서 건너뛴 매체
            
            # 활성화된 각 매체에 대해 처리
            for media_display in active_media:
                # 매체가 SA인지 DA인지 판단하여 적절한 키 가져오기
                if media_display in available_sa:
                    media_key = get_media_key(media_display, 'SA')
                elif media_display in available_da:
                    media_key = get_media_key(media_display, 'DA')
                else:
                    continue
                
                # 캐싱된 함수로 벤치마크 조회
                benchmark = get_media_benchmarks(industry, media_key)
                if not benchmark:
                    skipped_media.append(media_display)
                    continue
                
                user_ratio = st.session_state.media_config.get(f"{media_display}_ratio", 0.0)
                
                # 매체 타입 판단
                if '_SA' in media_key:
                    category = '검색광고'
                    revenue_per_cv = 100000
                    media_name = media_key.split('_')[0]
                else:
                    category = '디스플레이광고'
                    revenue_per_cv = 80000
                    media_name = media_key.replace('_DA', '').replace('DA', '')
                
                # 성과 지표 계산
                adjusted_cpc = apply_budget_adjustment(benchmark['cpc'], budget)
                adjusted_ctr = benchmark['ctr'] * SEASONALITY[month]
                adjusted_cvr = benchmark['cvr'] * SEASONALITY[month]
                
                media = {
                    'name': media_name,
                    'category': category,
                    'budget_ratio': user_ratio,
                    'cpc': adjusted_cpc,
                    'ctr': round(adjusted_ctr, 2),
                    'cvr': round(adjusted_cvr, 2),
                    'revenue_per_cv': revenue_per_cv,
                    'adjustment': 0
                }
                selected_media.append(media)
            
            # 건너뛴 매체 경고
            if skipped_media:
                st.warning(f"⚠️ 다음 매체는 '{industry}' 업종 벤치마크 데이터가 없어 제외되었습니다: {', '.join(skipped_media)}")
            
            # 선택된 매체 확인
            if not selected_media:
                st.error("❌ 선택한 매체 중 시뮬레이션 가능한 매체가 없습니다. 다른 매체를 선택해주세요.")
                raise ValueError("No valid media selected")
            
            # 시나리오 생성
            scenarios_data = {
                'conservative': [],
                'base': [],
                'aggressive': []
            }
            
            for media in selected_media:
                conservative_media = media.copy()
                conservative_media['adjustment'] = -5
                scenarios_data['conservative'].append(conservative_media)
                
                base_media = media.copy()
                base_media['adjustment'] = 0
                scenarios_data['base'].append(base_media)
                
                aggressive_media = media.copy()
                aggressive_media['adjustment'] = 10
                scenarios_data['aggressive'].append(aggressive_media)
            
            # 성과 계산
            scenarios = {}
            for scenario_key, media_list in scenarios_data.items():
                scenarios[scenario_key] = []
                for media in media_list:
                    media_performance = calculate_media_performance(media, budget)
                    scenarios[scenario_key].append(media_performance)
            
            # 결과 저장
            st.session_state.results = {
                'scenarios': scenarios,
                'budget': budget,
                'selected_media': selected_media,
                'industry': industry,
                'goal': goal_name,
                'month': month
            }
            
            status_text.text("✅ 완료!")
            progress_bar.progress(100)
            time.sleep(0.5)
            
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.success(f"✅ AI가 {industry} 업종 기준으로 최적 미디어믹스를 생성했습니다!")
            
        except ZeroDivisionError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 계산 오류: 0으로 나누기")
            st.warning("""
            **문제 원인:**
            - CPC, CTR, CVR 값이 0이거나 너무 작습니다.
            - 예산 배분이 잘못되었습니다.
            
            **해결 방법:**
            1. 매체별 비중을 확인하세요 (모든 매체가 0%가 아닌지)
            2. 예산이 충분한지 확인하세요 (최소 100만원 이상 권장)
            3. 선택된 매체가 있는지 확인하세요
            """)
            
            with st.expander("🔍 상세 오류 정보 (개발자용)"):
                st.code(str(e))
        
        except ValueError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 입력값 오류")
            st.warning(f"""
            **문제 원인:**
            잘못된 값이 입력되었습니다.
            
            **해결 방법:**
            1. 모든 입력값이 양수인지 확인하세요
            2. 매체 비중의 합이 100%인지 확인하세요
            3. 업종과 운영 월이 선택되었는지 확인하세요
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
        
        except KeyError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 데이터 누락 오류")
            st.warning(f"""
            **문제 원인:**
            필수 데이터가 누락되었습니다: {e}
            
            **해결 방법:**
            1. 매체를 다시 선택해주세요
            2. 페이지를 새로고침하세요 (Ctrl + R)
            3. 업종 벤치마크 데이터를 확인하세요
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
        
        except Exception as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("⚠️ 예기치 않은 오류가 발생했습니다")
            st.warning("""
            **해결 방법:**
            1. 입력값을 다시 확인해주세요
            2. 페이지를 새로고침하세요 (Ctrl + R)
            3. 문제가 계속되면 개발팀에 문의하세요
            
            **연락처:** JH Performance Marketing 10 Team
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
                import traceback
                st.code(traceback.format_exc())
    
    # 프리셋 저장 섹션
    st.markdown("---")
    with st.expander("💾 현재 설정을 프리셋으로 저장"):
        preset_name = st.text_input(
            "프리셋 이름",
            placeholder="예: 삼성생명_보험상품_Q1",
            key="save_preset_name"
        )
        
        if st.button("프리셋 저장", key="save_preset_btn"):
            if preset_name:
                try:
                    preset_data = {
                        "mode": "AI",
                        "budget": budget,
                        "industry": industry,
                        "month": month,
                        "goal": goal,
                        "media_config": st.session_state.media_config,
                        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    with open(f'saved_presets/{preset_name}.json', 'w', encoding='utf-8') as f:
                        json.dump(preset_data, f, ensure_ascii=False, indent=2)
                    
                    st.success(f"✅ '{preset_name}' 프리셋 저장 완료!")
                    # 프리셋 불러온 것 초기화
                    if 'preset_budget' in st.session_state:
                        del st.session_state.preset_budget
                    if 'preset_industry' in st.session_state:
                        del st.session_state.preset_industry
                    if 'preset_month' in st.session_state:
                        del st.session_state.preset_month
                    if 'preset_goal' in st.session_state:
                        del st.session_state.preset_goal
                    if 'preset_media_config' in st.session_state:
                        del st.session_state.preset_media_config
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 저장 실패: {e}")
            else:
                st.warning("⚠️ 프리셋 이름을 입력하세요")


# ===== 수동 입력 모드 =====
elif mode == "✏️ 수동 입력":
    # 공통 헤더
    render_page_header("수동 입력")
    st.markdown("매체별 데이터를 직접 입력하여 시뮬레이션합니다.")
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        # 프리셋 예산 적용
        default_manual_budget = st.session_state.get('preset_budget', 10000000)
        
        budget = st.number_input(
            "💰 총 예산 (원)",
            min_value=1000000,
            max_value=2000000000,
            value=default_manual_budget,
            step=1000000,
            format="%d",
            key="manual_budget_widget",
            help="""
            📌 Tips:
            • 1천만원 이상 권장
            • 예산이 클수록 CPC 경쟁도 상승
            • 최대 20억원까지 입력 가능
            """
        )
        st.session_state.budget = budget
        
        # 프리셋 매체 데이터가 있으면 개수 설정
        preset_media_data = st.session_state.get('preset_media_data', [])
        default_num_media = len(preset_media_data) if preset_media_data else 2
        
        st.markdown("---")
        num_media = st.number_input(
            "매체 개수",
            min_value=1,
            max_value=10,
            value=default_num_media,
            step=1
        )
    
    with col2:
        st.info(f"💰 총 예산: **{format_number(budget)}원**")
        st.info(f"📱 매체 수: **{num_media}개**")
    
    st.markdown("---")
    st.subheader("📝 매체별 데이터 입력")
    
    # 검증 우회 옵션
    col_check1, col_check2 = st.columns([3, 1])
    with col_check1:
        skip_validation = st.checkbox(
            "⚠️ 효율 검증 건너뛰기 (고급 사용자용)",
            help="업종 평균과 다른 특수한 케이스에 체크하세요",
            key="skip_validation"
        )
    
    selected_media = []
    total_ratio = 0
    
    # 프리셋 알림 표시
    if preset_media_data:
        st.info("💡 불러온 프리셋 데이터가 적용되었습니다. 수정 후 저장하거나 그대로 사용할 수 있습니다.")
    
    for i in range(num_media):
        st.markdown(f"### 매체 {i+1}")
        col1, col2, col3 = st.columns(3)
        
        # 프리셋 데이터가 있으면 사용
        preset_media = preset_media_data[i] if i < len(preset_media_data) else {}
        
        with col1:
            name = st.text_input(
                f"매체명", 
                value=preset_media.get('name', f"매체{i+1}"), 
                key=f"name_{i}"
            )
            
            categories = ["검색광고", "디스플레이광고", "동영상광고"]
            cat_index = categories.index(preset_media.get('category', '검색광고')) if preset_media.get('category') in categories else 0
            category = st.selectbox(
                f"카테고리",
                categories,
                index=cat_index,
                key=f"category_{i}"
            )
            
            # 업종 선택 (검증용)
            industry_list = list(EFFICIENCY_RANGES.keys())
            industry_for_validation = st.selectbox(
                f"업종 (검증용)",
                industry_list,
                key=f"industry_val_{i}",
                help="효율 검증을 위한 업종 선택"
            )
            
            budget_ratio = st.number_input(
                f"예산 비중 (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(preset_media.get('budget_ratio', 100.0/num_media)),
                step=0.1,
                key=f"ratio_{i}"
            )
        
        with col2:
            cpc = st.number_input(
                f"CPC (원)", 
                min_value=1, 
                value=int(preset_media.get('cpc', 500)), 
                step=50, 
                key=f"cpc_{i}",
                help="클릭당 비용 (Cost Per Click)"
            )
            # CPC 검증
            if not skip_validation:
                warning = validate_efficiency('CPC', cpc, industry_for_validation)
                if warning:
                    st.caption(warning)
            
            ctr = st.number_input(
                f"CTR (%)", 
                min_value=0.01, 
                value=float(preset_media.get('ctr', 2.0)), 
                step=0.1, 
                key=f"ctr_{i}",
                help="클릭률 (Click Through Rate)"
            )
            # CTR 검증
            if not skip_validation:
                warning = validate_efficiency('CTR', ctr, industry_for_validation)
                if warning:
                    st.caption(warning)
            
            cvr = st.number_input(
                f"CVR (%)", 
                min_value=0.01, 
                value=float(preset_media.get('cvr', 2.0)), 
                step=0.1, 
                key=f"cvr_{i}",
                help="전환율 (Conversion Rate)"
            )
            # CVR 검증
            if not skip_validation:
                warning = validate_efficiency('CVR', cvr, industry_for_validation)
                if warning:
                    st.caption(warning)
        
        with col3:
            revenue = st.number_input(
                f"전환당 매출 (원)",
                min_value=1000,
                value=int(preset_media.get('revenue_per_cv', 100000)),
                step=10000,
                key=f"revenue_{i}"
            )
            adjustment = st.number_input(
                f"예측 오차 (%)",
                min_value=-50.0,
                max_value=50.0,
                value=float(preset_media.get('adjustment', 0.0)),
                step=1.0,
                key=f"adjustment_{i}"
            )
        
        media = {
            'name': name,
            'category': category,
            'budget_ratio': budget_ratio,
            'cpc': cpc,
            'ctr': ctr,
            'cvr': cvr,
            'revenue_per_cv': revenue,
            'adjustment': adjustment
        }
        selected_media.append(media)
        total_ratio += budget_ratio
        
        st.markdown("---")
    
    # 사이드바에 실시간 합계 표시
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 입력 현황")
    
    if total_ratio < 100:
        st.sidebar.warning(f"예산 비중 합계: {total_ratio:.1f}%\n남은 비중: {100 - total_ratio:.1f}%")
    elif abs(total_ratio - 100) < 0.01:
        st.sidebar.success(f"예산 비중 합계: {total_ratio:.1f}% ✅")
    else:
        st.sidebar.error(f"예산 비중 합계: {total_ratio:.1f}%\n초과: {total_ratio - 100:.1f}%")
    
    # 진행률 바
    st.sidebar.progress(min(total_ratio / 100, 1.0))
    
    # 예산 비중 합계 표시
    if abs(total_ratio - 100) < 0.01:
        st.success(f"✅ 예산 비중 합계: **{total_ratio:.1f}%** (정상)")
    else:
        st.warning(f"⚠️ 예산 비중 합계: **{total_ratio:.1f}%** (100%가 되어야 합니다)")
    
    if st.button("🚀 계산 실행", type="primary", use_container_width=True):
        if abs(total_ratio - 100) > 0.01:
            st.error("❌ 예산 비중 합계가 100%가 되어야 합니다.")
        else:
            # 로딩 애니메이션
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                status_text.text("📊 매체 데이터 검증 중...")
                progress_bar.progress(25)
                time.sleep(0.3)
                
                status_text.text("💰 예산 배분 계산...")
                progress_bar.progress(50)
                time.sleep(0.3)
                
                status_text.text("📈 성과 예측 중...")
                progress_bar.progress(75)
                
                # 시나리오 생성 (고정: ±5%)
                scenario_adjustment = 5.0
                scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
                
                st.session_state.results = {
                    'scenarios': scenarios,
                    'budget': budget,
                    'selected_media': selected_media,
                    'scenario_adjustment': scenario_adjustment,
                    'industry': '보험',  # 기본값
                    'month': 1,  # 기본값
                    'goal': '균형'  # 기본값
                }
                
                status_text.text("✅ 완료!")
                progress_bar.progress(100)
                time.sleep(0.5)
                
                # 정리
                progress_bar.empty()
                status_text.empty()
                
                st.success("✅ 계산이 완료되었습니다!")
                
            except ZeroDivisionError as e:
                # 정리
                progress_bar.empty()
                status_text.empty()
                
                st.error("❌ 계산 오류: 0으로 나누기")
                st.warning("""
                **문제 원인:**
                - CPC, CTR, CVR 값이 0이거나 너무 작습니다.
                - 예산 배분이 잘못되었습니다.
                
                **해결 방법:**
                1. 각 매체의 CPC, CTR, CVR 값을 확인하세요 (0보다 커야 함)
                2. 예산이 충분한지 확인하세요
                3. 비중 합계가 100%인지 확인하세요
                """)
                
                with st.expander("🔍 상세 오류 정보"):
                    st.code(str(e))
            
            except ValueError as e:
                # 정리
                progress_bar.empty()
                status_text.empty()
                
                st.error("❌ 입력값 오류")
                st.warning("""
                **문제 원인:**
                잘못된 값이 입력되었습니다.
                
                **해결 방법:**
                1. 모든 매체의 CPC, CTR, CVR이 양수인지 확인하세요
                2. 예산 비중이 모두 0 이상인지 확인하세요
                3. 전환당 매출액이 입력되었는지 확인하세요
                """)
                
                with st.expander("🔍 상세 오류 정보"):
                    st.code(str(e))
            
            except Exception as e:
                # 정리
                progress_bar.empty()
                status_text.empty()
                
                st.error("⚠️ 계산 중 오류가 발생했습니다")
                
                with st.expander("🔍 상세 오류 정보"):
                    st.code(str(e))
                
                st.markdown("""
                **해결 방법:**
                1. 모든 입력값이 올바른지 확인하세요
                2. 예산 비중 합계가 100%인지 확인하세요
                3. 페이지를 새로고침하세요 (Ctrl + R)
                
                **연락처:** JH Performance Marketing 10 Team
                """)
    
    # 프리셋 저장 섹션
    st.markdown("---")
    with st.expander("💾 현재 설정을 프리셋으로 저장"):
        preset_name_manual = st.text_input(
            "프리셋 이름",
            placeholder="예: 브라보코리아_외국인타겟_12월",
            key="save_manual_preset_name"
        )
        
        if st.button("프리셋 저장", key="save_manual_preset_btn"):
            if preset_name_manual:
                try:
                    preset_data = {
                        "mode": "Manual",
                        "budget": budget,
                        "media_data": selected_media,
                        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    with open(f'saved_presets/{preset_name_manual}.json', 'w', encoding='utf-8') as f:
                        json.dump(preset_data, f, ensure_ascii=False, indent=2)
                    
                    st.success(f"✅ '{preset_name_manual}' 프리셋 저장 완료!")
                    # 프리셋 불러온 것 초기화
                    if 'preset_budget' in st.session_state:
                        del st.session_state.preset_budget
                    if 'preset_media_data' in st.session_state:
                        del st.session_state.preset_media_data
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 저장 실패: {e}")
            else:
                st.warning("⚠️ 프리셋 이름을 입력하세요")


# ===== 목표 역산 계산 모드 =====
elif mode == "🎯 목표 역산 계산":
    # 공통 헤더
    render_page_header("목표 역산 계산")
    
    st.markdown("""
    ### 목표 달성에 필요한 예산을 계산합니다
    원하는 전환수와 목표 CPA를 입력하면, 필요한 총 예산과 매체별 배분을 추천합니다.
    """)
    
    st.markdown("---")
    
    # 기본 정보 입력
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📊 목표 설정")
        
        target_conversions = st.number_input(
            "🎯 목표 전환수 (건)",
            min_value=1,
            value=1000,
            step=10,
            key="target_cv_widget",
            help="달성하고자 하는 월간 전환수 목표"
        )
        
        target_cpa = st.number_input(
            "💰 목표 CPA (원)",
            min_value=1000,
            value=50000,
            step=1000,
            key="target_cpa_widget",
            help="전환 1건당 목표 비용 (Cost Per Action)"
        )
        
        # 계산된 필요 예산
        required_budget = target_conversions * target_cpa
        
        st.metric(
            "📈 필요한 총 예산",
            f"{required_budget:,}원",
            help="목표 전환수 × 목표 CPA"
        )
    
    with col2:
        st.markdown("#### 🏢 업종 정보")
        
        industry_list = list(INDUSTRY_BASE_METRICS.keys())
        industry = st.selectbox(
            "업종 선택",
            industry_list,
            key="target_industry"
        )
        
        month = st.selectbox(
            "운영 월",
            list(range(1, 13)),
            format_func=lambda x: f"{x}월",
            key="target_month"
        )
        
        goal = st.radio(
            "캠페인 목표",
            ["인지도 중심", "전환 중심", "균형"],
            index=1,  # 기본값: 전환 중심
            key="target_goal"
        )
    
    st.markdown("---")
    
    # 고급 옵션
    with st.expander("🔧 고급 설정 (선택)"):
        col1, col2 = st.columns(2)
        
        with col1:
            # 매체 제외 옵션
            exclude_media = st.multiselect(
                "제외할 매체",
                ALL_MEDIA,
                key="target_exclude"
            )
        
        with col2:
            confidence_level = st.slider(
                "예측 신뢰도",
                0.5, 1.5, 1.0, 0.1,
                help="1.0 = 기준, 0.5 = 보수적, 1.5 = 공격적",
                key="target_confidence"
            )
    
    st.markdown("---")
    
    # 계산 실행 버튼
    if st.button("🚀 목표 역산 계산 실행", type="primary", use_container_width=True, key="run_target_calc"):
        # 로딩 애니메이션
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("📊 목표 분석 중...")
            progress_bar.progress(20)
            time.sleep(0.3)
            
            status_text.text("💰 필요 예산 계산...")
            progress_bar.progress(40)
            time.sleep(0.3)
            
            status_text.text("📈 매체별 배분 계산...")
            progress_bar.progress(60)
            time.sleep(0.3)
            
            status_text.text("🎯 달성 가능성 분석...")
            progress_bar.progress(80)
            
            # 목표 역산 계산 로직
            
            # 1. 업종 기본 지표 가져오기
            base_metrics = INDUSTRY_BASE_METRICS.get(industry, {})
            
            # 2. 목표별 예산 배분 비율
            if goal == "인지도 중심":
                allocation = {"DA": 0.6, "SA": 0.4}
            elif goal == "전환 중심":
                allocation = {"SA": 0.7, "DA": 0.3}
            else:  # 균형
                allocation = {"SA": 0.5, "DA": 0.5}
            
            # 3. 매체 목록 매핑
            media_display_map = {
                '네이버 검색광고': '네이버_SA',
                '구글 검색광고': '구글_SA',
                '카카오 검색광고': '카카오_SA',
                '메타': '메타_DA',
                '구글 배너광고': '구글GDA_DA',
                '틱톡': '틱톡_DA',
                '네이버 GFA': '네이버GFA_DA',
                '카카오 모먼트': '카카오모먼트_DA'
            }
            
            # 4. 매체별 세부 배분
            media_list = []
            
            # SA 매체
            sa_budget = required_budget * allocation["SA"]
            sa_media_display = [m for m in MEDIA_CATEGORIES['SA'] if m not in exclude_media]
            
            if sa_media_display:
                sa_budget_per_media = sa_budget / len(sa_media_display)
                
                for media_display in sa_media_display:
                    media_key = media_display_map[media_display]
                    
                    # 매체별 보정된 지표
                    adjusted_metrics = get_media_adjusted_metrics(
                        industry, media_key, month, sa_budget_per_media
                    )
                    
                    if adjusted_metrics:
                        # 성과 계산
                        performance = calculate_performance(sa_budget_per_media, adjusted_metrics)
                        
                        media_list.append({
                            'name': media_display,
                            'category': '검색광고',
                            'budget': sa_budget_per_media,
                            'ratio': (sa_budget_per_media / required_budget) * 100,
                            **performance
                        })
            
            # DA 매체
            da_budget = required_budget * allocation["DA"]
            da_media_display = [m for m in MEDIA_CATEGORIES['DA'] if m not in exclude_media]
            
            if da_media_display:
                da_budget_per_media = da_budget / len(da_media_display)
                
                for media_display in da_media_display:
                    media_key = media_display_map[media_display]
                    
                    # 매체별 보정된 지표
                    adjusted_metrics = get_media_adjusted_metrics(
                        industry, media_key, month, da_budget_per_media
                    )
                    
                    if adjusted_metrics:
                        performance = calculate_performance(da_budget_per_media, adjusted_metrics)
                        
                        media_list.append({
                            'name': media_display,
                            'category': '디스플레이광고',
                            'budget': da_budget_per_media,
                            'ratio': (da_budget_per_media / required_budget) * 100,
                            **performance
                        })
            
            # 5. 전체 성과 집계
            total_conversions = sum(m['conversions'] for m in media_list)
            avg_cpa = required_budget / total_conversions if total_conversions > 0 else 0
            
            # 6. 신뢰도 보정 적용
            predicted_conversions = total_conversions * confidence_level
            predicted_cpa = avg_cpa / confidence_level if confidence_level > 0 else avg_cpa
            achievement_rate = (predicted_conversions / target_conversions) * 100
            
            # 7. 결과 저장
            result = {
                'required_budget': required_budget,
                'target_conversions': target_conversions,
                'target_cpa': target_cpa,
                'predicted_conversions': predicted_conversions,
                'predicted_cpa': predicted_cpa,
                'achievement_rate': achievement_rate,
                'media_list': media_list,
                'industry': industry,
                'month': month,
                'goal': goal
            }
            
            status_text.text("✅ 완료!")
            progress_bar.progress(100)
            time.sleep(0.5)
            
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.success("✅ 계산 완료!")
            
            # === 결과 출력 ===
            st.markdown("---")
            st.markdown("### 📊 목표 달성 예측")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "필요 예산",
                    f"{result['required_budget']:,}원"
                )
            
            with col2:
                st.metric(
                    "목표 전환수",
                    f"{result['target_conversions']:,}건"
                )
            
            with col3:
                delta = result['predicted_conversions'] - result['target_conversions']
                st.metric(
                    "예상 전환수",
                    f"{result['predicted_conversions']:,.0f}건",
                    f"{delta:+,.0f}건"
                )
            
            with col4:
                achievement = result['achievement_rate']
                st.metric(
                    "달성률",
                    f"{achievement:.1f}%"
                )
            
            # 달성 가능성 평가
            if achievement >= 100:
                st.success("✅ 목표 달성 가능성이 높습니다!")
            elif achievement >= 80:
                st.warning("⚠️ 목표 달성을 위해 효율 개선이 필요합니다.")
            else:
                st.error("❌ 현재 조건으로는 목표 달성이 어렵습니다. 예산 증액 또는 목표 조정을 권장합니다.")
            
            st.markdown("---")
            
            # 매체별 배분 테이블
            st.markdown("### 📋 매체별 예산 배분 및 예상 성과")
            
            if media_list:
                df = pd.DataFrame(media_list)
                
                # 컬럼 포맷팅
                display_df = df.copy()
                display_df['budget'] = display_df['budget'].apply(lambda x: f"{x:,.0f}원")
                display_df['ratio'] = display_df['ratio'].apply(lambda x: f"{x:.1f}%")
                display_df['impressions'] = display_df['impressions'].apply(lambda x: f"{x:,.0f}")
                display_df['clicks'] = display_df['clicks'].apply(lambda x: f"{x:,.0f}")
                display_df['conversions'] = display_df['conversions'].apply(lambda x: f"{x:,.1f}")
                display_df['cpa'] = display_df['cpa'].apply(lambda x: f"{x:,.0f}원")
                display_df['ctr'] = display_df['ctr'].apply(lambda x: f"{x:.2f}%")
                display_df['cvr'] = display_df['cvr'].apply(lambda x: f"{x:.2f}%")
                display_df['cpc'] = display_df['cpc'].apply(lambda x: f"{x:,.0f}원")
                
                # 컬럼명 한글화
                display_df.rename(columns={
                    'name': '매체명',
                    'category': '카테고리',
                    'budget': '예산',
                    'ratio': '비중',
                    'impressions': '노출수',
                    'clicks': '클릭수',
                    'conversions': '전환수',
                    'cpa': 'CPA',
                    'ctr': 'CTR',
                    'cvr': 'CVR',
                    'cpc': 'CPC'
                }, inplace=True)
                
                st.dataframe(display_df, use_container_width=True, hide_index=True)
                
                # 차트
                col1, col2 = st.columns(2)
                
                with col1:
                    # 예산 비중 파이차트
                    fig_budget = px.pie(
                        df, 
                        values='budget', 
                        names='name',
                        title='매체별 예산 비중'
                    )
                    st.plotly_chart(fig_budget, use_container_width=True)
                
                with col2:
                    # 전환수 바차트
                    fig_cv = px.bar(
                        df,
                        x='name',
                        y='conversions',
                        title='매체별 예상 전환수',
                        labels={'name': '매체', 'conversions': '전환수'}
                    )
                    st.plotly_chart(fig_cv, use_container_width=True)
                
                # Excel 다운로드
                st.markdown("---")
                st.markdown("### 📥 결과 다운로드")
                
                # Excel 생성 - 목표 역산은 시나리오가 없으므로 별도 처리 (최적화 버전)
                from openpyxl.styles import Alignment, Border, Side
                
                # 테두리 스타일
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # 요약 시트
                    summary_data = {
                        '항목': ['필요 예산', '목표 전환수', '목표 CPA', '예상 전환수', '예상 CPA', '달성률', '업종', '운영 월', '캠페인 목표'],
                        '값': [
                            f"{result['required_budget']:,}원",
                            f"{result['target_conversions']:,}건",
                            f"{result['target_cpa']:,}원",
                            f"{result['predicted_conversions']:,.1f}건",
                            f"{result['predicted_cpa']:,.0f}원",
                            f"{result['achievement_rate']:.1f}%",
                            result['industry'],
                            f"{result['month']}월",
                            result['goal']
                        ]
                    }
                    summary_df_target = pd.DataFrame(summary_data)
                    summary_df_target.to_excel(writer, sheet_name='요약', index=False)
                    
                    # 요약 시트 포맷 (최적화)
                    ws_summary = writer.sheets['요약']
                    num_rows_summary = len(summary_df_target) + 1
                    
                    # 열 너비 설정
                    for col_idx in range(1, 3):  # 항목, 값
                        col_letter = ws_summary.cell(row=1, column=col_idx).column_letter
                        ws_summary.column_dimensions[col_letter].width = 15
                    
                    # 가운데 정렬
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    for row in range(1, num_rows_summary + 1):
                        for col in range(1, 3):
                            ws_summary.cell(row=row, column=col).alignment = center_alignment
                    
                    # 테두리: 헤더 + 마지막 행만
                    for col in range(1, 3):
                        ws_summary.cell(row=1, column=col).border = thin_border
                        ws_summary.cell(row=num_rows_summary, column=col).border = thin_border
                    
                    # 매체별 상세 시트 - DataFrame 단계에서 퍼센티지 변환
                    df_detail = df.copy()
                    percentage_columns = ['예산비중', 'CTR', 'CVR', 'ROAS']
                    for col in percentage_columns:
                        if col in df_detail.columns:
                            df_detail[col] = df_detail[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) and x != '' else x)
                    
                    df_detail.to_excel(writer, sheet_name='매체별 상세', index=False)
                    
                    # 상세 시트 포맷 (최적화)
                    ws_detail = writer.sheets['매체별 상세']
                    num_rows_detail = len(df_detail) + 1
                    num_cols_detail = len(df_detail.columns)
                    
                    # 열 너비 설정
                    for col_idx in range(1, num_cols_detail + 1):
                        col_letter = ws_detail.cell(row=1, column=col_idx).column_letter
                        ws_detail.column_dimensions[col_letter].width = 15
                    
                    # 가운데 정렬
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    for row in range(1, num_rows_detail + 1):
                        for col in range(1, num_cols_detail + 1):
                            ws_detail.cell(row=row, column=col).alignment = center_alignment
                    
                    # 테두리: 헤더 + 마지막 행만
                    for col in range(1, num_cols_detail + 1):
                        ws_detail.cell(row=1, column=col).border = thin_border
                        ws_detail.cell(row=num_rows_detail, column=col).border = thin_border
                
                excel_data = output.getvalue()
                
                now = datetime.now()
                filename = f"목표역산_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                st.download_button(
                    label="📥 Excel 다운로드",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        except ZeroDivisionError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 계산 오류: 0으로 나누기")
            st.warning("""
            **문제 원인:**
            - 목표 CPA 또는 목표 전환수가 0이거나 너무 작습니다.
            - 업종 벤치마크 데이터에 문제가 있습니다.
            
            **해결 방법:**
            1. 목표 전환수를 1 이상으로 설정하세요
            2. 목표 CPA를 1,000원 이상으로 설정하세요
            3. 다른 업종을 선택해보세요
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
        
        except ValueError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 입력값 오류")
            st.warning("""
            **문제 원인:**
            잘못된 값이 입력되었습니다.
            
            **해결 방법:**
            1. 목표 전환수는 양수여야 합니다 (1건 이상)
            2. 목표 CPA는 양수여야 합니다 (1,000원 이상 권장)
            3. 업종과 운영 월을 올바르게 선택하세요
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
        
        except KeyError as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("❌ 데이터 누락 오류")
            st.warning(f"""
            **문제 원인:**
            필수 벤치마크 데이터가 누락되었습니다: {e}
            
            **해결 방법:**
            1. 다른 업종을 선택해보세요
            2. 선택한 매체가 해당 업종에 지원되는지 확인하세요
            3. 페이지를 새로고침하세요 (Ctrl + R)
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
        
        except Exception as e:
            # 정리
            progress_bar.empty()
            status_text.empty()
            
            st.error("⚠️ 계산 중 오류가 발생했습니다")
            st.warning("""
            **해결 방법:**
            1. 목표 전환수와 목표 CPA를 다시 확인하세요
            2. 업종과 운영 월이 올바른지 확인하세요
            3. 페이지를 새로고침하세요 (Ctrl + R)
            
            **연락처:** JH Performance Marketing 10 Team
            """)
            
            with st.expander("🔍 상세 오류 정보"):
                st.code(str(e))
                import traceback
                st.code(traceback.format_exc())


# ===== 엑셀 업로드 모드 =====
elif mode == "📁 엑셀 업로드":
    # 공통 헤더
    render_page_header("엑셀 업로드")
    st.markdown("미디어믹스 입력 템플릿을 업로드하여 시뮬레이션합니다.")
    
    uploaded_file = st.file_uploader(
        "Excel 파일 선택 (.xlsx)",
        type=['xlsx'],
        help="미디어믹스_입력.xlsx 템플릿을 사용하세요."
    )
    
    if uploaded_file is not None:
        try:
            from openpyxl import load_workbook
            
            # 파일 읽기
            wb = load_workbook(uploaded_file)
            ws = wb.active
            
            # 총 예산 읽기
            budget = ws['B1'].value
            
            if budget is None or budget <= 0:
                st.error("❌ 총 예산이 올바르지 않습니다. (B1 셀 확인)")
            else:
                st.success(f"✅ 총 예산: **{format_number(budget)}원**")
                
                # 카테고리 매핑
                category_map = {
                    "SA": "검색광고",
                    "DA": "디스플레이광고",
                    "VA": "동영상광고"
                }
                
                # 데이터 읽기
                selected_media = []
                row = 4
                
                while True:
                    media_name = ws.cell(row=row, column=1).value
                    if media_name is None or str(media_name).strip() == "":
                        break
                    
                    category_code = ws.cell(row=row, column=2).value
                    budget_ratio_raw = ws.cell(row=row, column=3).value
                    cpc = ws.cell(row=row, column=4).value
                    ctr_raw = ws.cell(row=row, column=5).value
                    cvr_raw = ws.cell(row=row, column=6).value
                    adjustment_raw = ws.cell(row=row, column=7).value
                    
                    # % 문자 제거
                    def parse_percentage(value):
                        if isinstance(value, str):
                            value = value.replace('%', '').strip()
                        try:
                            return float(value)
                        except:
                            return 0
                    
                    budget_ratio = parse_percentage(budget_ratio_raw)
                    ctr = parse_percentage(ctr_raw)
                    cvr = parse_percentage(cvr_raw)
                    adjustment = parse_percentage(adjustment_raw)
                    
                    category = category_map.get(str(category_code).upper(), "검색광고")
                    
                    media = {
                        'name': str(media_name).strip(),
                        'category': category,
                        'budget_ratio': budget_ratio,
                        'cpc': int(cpc) if cpc else 0,
                        'ctr': ctr,
                        'cvr': cvr,
                        'revenue_per_cv': 100000,
                        'adjustment': adjustment
                    }
                    
                    selected_media.append(media)
                    row += 1
                
                if selected_media:
                    # 데이터 미리보기
                    st.subheader("📊 데이터 미리보기")
                    preview_data = []
                    for media in selected_media:
                        preview_data.append({
                            '매체명': media['name'],
                            '카테고리': media['category'],
                            '예산비중': f"{media['budget_ratio']:.1f}%",
                            'CPC': f"{media['cpc']:,}원",
                            'CTR': f"{media['ctr']:.1f}%",
                            'CVR': f"{media['cvr']:.1f}%",
                            '예측오차': f"{media['adjustment']:+.1f}%"
                        })
                    
                    df_preview = pd.DataFrame(preview_data)
                    st.dataframe(df_preview, use_container_width=True)
                    
                    total_ratio = sum(m['budget_ratio'] for m in selected_media)
                    if abs(total_ratio - 100) < 0.01:
                        st.success(f"✅ 예산 비중 합계: **{total_ratio:.1f}%**")
                    else:
                        st.warning(f"⚠️ 예산 비중 합계: **{total_ratio:.1f}%** (100%가 아닙니다)")
                    
                    if st.button("🚀 계산 실행", type="primary", use_container_width=True):
                        # 로딩 애니메이션
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        try:
                            status_text.text("📊 Excel 데이터 분석 중...")
                            progress_bar.progress(25)
                            time.sleep(0.3)
                            
                            status_text.text("💰 예산 배분 계산...")
                            progress_bar.progress(50)
                            time.sleep(0.3)
                            
                            status_text.text("📈 성과 예측 중...")
                            progress_bar.progress(75)
                            
                            scenario_adjustment = 5.0
                            scenarios = generate_scenarios(selected_media, budget, scenario_adjustment)
                            
                            st.session_state.results = {
                                'scenarios': scenarios,
                                'budget': budget,
                                'selected_media': selected_media,
                                'scenario_adjustment': scenario_adjustment,
                                'industry': '보험',  # 기본값
                                'month': 1,  # 기본값
                                'goal': '균형'  # 기본값
                            }
                            
                            status_text.text("✅ 완료!")
                            progress_bar.progress(100)
                            time.sleep(0.5)
                            
                            # 정리
                            progress_bar.empty()
                            status_text.empty()
                            
                            st.success("✅ 계산이 완료되었습니다!")
                            
                        except ZeroDivisionError as e:
                            # 정리
                            progress_bar.empty()
                            status_text.empty()
                            
                            st.error("❌ 계산 오류: 0으로 나누기")
                            st.warning("""
                            **문제 원인:**
                            - Excel 파일에 CPC, CTR, CVR 값이 0이거나 누락되었습니다.
                            
                            **해결 방법:**
                            1. Excel 파일에서 CPC, CTR, CVR 값을 확인하세요 (모두 0보다 커야 함)
                            2. 예산 비중이 올바르게 입력되었는지 확인하세요
                            3. 템플릿을 다시 다운로드하여 작성하세요
                            """)
                            
                            with st.expander("🔍 상세 오류 정보"):
                                st.code(str(e))
                        
                        except ValueError as e:
                            # 정리
                            progress_bar.empty()
                            status_text.empty()
                            
                            st.error("❌ 입력값 오류")
                            st.warning("""
                            **문제 원인:**
                            Excel 파일에 잘못된 값이 포함되어 있습니다.
                            
                            **해결 방법:**
                            1. 모든 숫자 값이 양수인지 확인하세요
                            2. 셀에 텍스트가 아닌 숫자가 입력되었는지 확인하세요
                            3. 예산 비중 합계가 100%인지 확인하세요
                            """)
                            
                            with st.expander("🔍 상세 오류 정보"):
                                st.code(str(e))
                        
                        except Exception as e:
                            # 정리
                            progress_bar.empty()
                            status_text.empty()
                            
                            st.error("⚠️ 계산 중 오류가 발생했습니다")
                            st.warning("""
                            **해결 방법:**
                            1. Excel 파일 형식을 다시 확인하세요
                            2. 템플릿을 다시 다운로드하여 사용하세요
                            3. 페이지를 새로고침하세요 (Ctrl + R)
                            
                            **연락처:** JH Performance Marketing 10 Team
                            """)
                            
                            with st.expander("🔍 상세 오류 정보"):
                                st.code(str(e))
                                import traceback
                                st.code(traceback.format_exc())
                else:
                    st.error("❌ 읽을 수 있는 매체 데이터가 없습니다.")
        
        except Exception as e:
            st.error(f"❌ 파일 읽기 오류: {e}")


# ===== 템플릿 다운로드 =====
elif mode == "📥 템플릿 다운로드":
    # 공통 헤더
    render_page_header("템플릿 다운로드")
    st.markdown("미디어믹스 데이터를 입력할 수 있는 템플릿을 다운로드합니다.")
    
    st.info("""
    ### 📋 템플릿 사용 방법
    1. 아래 버튼을 클릭하여 템플릿을 다운로드합니다.
    2. Excel에서 파일을 열고 데이터를 수정합니다.
    3. '엑셀 업로드' 모드에서 수정한 파일을 업로드합니다.
    
    ### 📊 입력 항목
    - **A1**: 총예산 (B1 셀에 숫자 입력)
    - **A4~**: 매체 데이터 (매체명, 카테고리, 예산비중, CPC, CTR, CVR, 예측오차)
    - **카테고리 코드**: SA (검색광고), DA (디스플레이광고), VA (동영상광고)
    """)
    
    # 템플릿 생성
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "총예산"
        ws['B1'] = 100000000
        
        headers = ["매체명", "카테고리", "예산비중", "평균 CPC", "평균 CTR", "평균 CVR", "예측 오차"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        sample_data = [
            ["네이버", "SA", "25%", 300, "3%", "4%", "1%"],
            ["구글", "SA", "25%", 400, "3%", "4%", "1%"],
            ["메타", "DA", "25%", 500, "3%", "4%", "10%"],
            ["카카오DA", "DA", "25%", 600, "3%", "4%", "-1%"]
        ]
        
        for row_idx, data in enumerate(sample_data, 4):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, 8):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        # 파일을 메모리에 저장
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        st.download_button(
            label="📥 템플릿 다운로드",
            data=buffer,
            file_name="미디어믹스_입력.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.success("✅ 템플릿이 준비되었습니다. 다운로드 버튼을 클릭하세요.")
    
    except Exception as e:
        st.error(f"❌ 템플릿 생성 오류: {e}")


# ===== 결과 출력 =====
if st.session_state.results is not None:
    st.markdown("---")
    st.header("📊 시뮬레이션 결과")
    
    results = st.session_state.results
    scenarios = results['scenarios']
    budget = results['budget']
    
    # 시나리오 요약
    st.subheader("🎯 시나리오 요약")
    
    summary_data = []
    scenario_names = {
        'conservative': '보수안(-5%)',
        'base': '기본안(0%)',
        'aggressive': '공격안(+10%)'
    }
    
    for scenario_key, scenario_name in scenario_names.items():
        scenario_data = scenarios[scenario_key]
        total_conversions = sum(m['estimated_conversions_adjusted'] for m in scenario_data)
        total_clicks = sum(m['estimated_clicks'] for m in scenario_data)
        avg_cpa = (budget / total_conversions) if total_conversions > 0 else 0
        total_revenue = sum(m['total_revenue_adjusted'] for m in scenario_data)
        avg_roas = (total_revenue / budget * 100) if budget > 0 else 0
        
        summary_data.append({
            '구분': scenario_name,
            '총전환수': int(total_conversions),
            '평균CPA': int(avg_cpa),
            '총클릭수': int(total_clicks),
            '평균ROAS': round(avg_roas, 1)
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # 핵심 지표 카드 (대형)
    st.markdown("### 📌 핵심 지표")
    col1, col2, col3, col4 = st.columns(4)
    base_data = summary_data[1]  # 기본안
    
    # 효율 등급 계산
    efficiency_grade = calculate_efficiency_grade(
        base_data['평균CPA'], 
        base_data['평균ROAS'], 
        base_data['총전환수']
    )
    
    grade_colors = {
        'S': '🏆',
        'A': '🥇', 
        'B': '🥈',
        'C': '🥉'
    }
    
    with col1:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #2196F3 0%, #1976D2 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;
                    box-shadow: 0 4px 12px rgba(33, 150, 243, 0.2);">
            <h3 style="color: white; margin: 0;">🎯 총 전환수</h3>
            <h1 style="color: white; margin: 10px 0; font-size: 2.5em;">{format_number(base_data['총전환수'])}</h1>
            <p style="color: rgba(255,255,255,0.9); margin: 0;">건</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #FF6B6B 0%, #EE5A6F 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;
                    box-shadow: 0 4px 12px rgba(255, 107, 107, 0.2);">
            <h3 style="color: white; margin: 0;">💵 평균 CPA</h3>
            <h1 style="color: white; margin: 10px 0; font-size: 2.5em;">{format_number(base_data['평균CPA'])}</h1>
            <p style="color: rgba(255,255,255,0.9); margin: 0;">원</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #4ECDC4 0%, #44A08D 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;
                    box-shadow: 0 4px 12px rgba(78, 205, 196, 0.2);">
            <h3 style="color: white; margin: 0;">📈 총 ROAS</h3>
            <h1 style="color: white; margin: 10px 0; font-size: 2.5em;">{base_data['평균ROAS']:.1f}</h1>
            <p style="color: rgba(255,255,255,0.9); margin: 0;">%</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #FFA726 0%, #FB8C00 100%); 
                    padding: 20px; border-radius: 10px; text-align: center;
                    box-shadow: 0 4px 12px rgba(255, 167, 38, 0.2);">
            <h3 style="color: white; margin: 0;">⭐ 효율 등급</h3>
            <h1 style="color: white; margin: 10px 0; font-size: 3em;">{efficiency_grade}</h1>
            <p style="color: rgba(255,255,255,0.9); margin: 0;">{grade_colors[efficiency_grade]} 등급</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("")  # 간격
    
    # 매체 순위 (CPA 기준)
    st.markdown("---")
    st.markdown("### 🏆 매체별 효율 순위")
    
    # base 시나리오의 매체 데이터 가져오기
    base_media = scenarios['base']
    media_ranking = []
    for media in base_media:
        cpa = media.get('cpa', 0)
        if cpa > 0:
            media_ranking.append({
                'name': media.get('name', ''),
                'cpa': cpa,
                'conversions': media.get('estimated_conversions_adjusted', 0),
                'budget_ratio': media.get('budget_ratio', 0),
                'roas': media.get('roas', 0)
            })
    
    # CPA 기준 정렬
    media_ranking.sort(key=lambda x: x['cpa'])
    
    rank_col1, rank_col2 = st.columns(2)
    
    with rank_col1:
        st.markdown("#### 🏆 Best 3 (CPA 낮은 순)")
        for i, media in enumerate(media_ranking[:3], 1):
            medal = ['🥇', '🥈', '🥉'][i-1]
            st.markdown(f"""
            <div style="background-color: #E3F2FD; padding: 15px; border-radius: 8px; margin-bottom: 10px; 
                        border-left: 4px solid #2196F3; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                <strong style="color: #1a1a1a;">{medal} {i}위: {media['name']}</strong><br>
                <span style="color: #424242;">CPA: <strong>{media['cpa']:,.0f}원</strong> | 
                전환: {media['conversions']:.0f}건 | 
                비중: {media['budget_ratio']:.1f}%</span>
            </div>
            """, unsafe_allow_html=True)
    
    with rank_col2:
        st.markdown("#### ⚠️ Worst 3 (CPA 높은 순)")
        worst_media = media_ranking[-3:][::-1]  # 뒤에서 3개, 역순
        for i, media in enumerate(worst_media, 1):
            st.markdown(f"""
            <div style="background-color: #FFF3E0; padding: 15px; border-radius: 8px; margin-bottom: 10px; 
                        border-left: 4px solid #FF9800; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                <strong style="color: #1a1a1a;">⚠️ {i}위: {media['name']}</strong><br>
                <span style="color: #424242;">CPA: <strong>{media['cpa']:,.0f}원</strong> | 
                전환: {media['conversions']:.0f}건 | 
                비중: {media['budget_ratio']:.1f}%</span>
            </div>
            """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # 시나리오 비교 테이블
    st.subheader("📊 시나리오 비교")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)
    
    # 차트
    st.markdown("---")
    st.subheader("📈 시각화")
    
    chart_col1, chart_col2 = st.columns(2)
    
    with chart_col1:
        st.markdown("##### 매체별 예산 비중")
        # 파이 차트 - 매체별 예산 비중
        selected_media = results.get('selected_media', [])
        if selected_media:
            pie_data = pd.DataFrame([{
                '매체명': media['name'],
                '예산비중': media['budget_ratio']
            } for media in selected_media])
            
            fig_pie = px.pie(
                pie_data,
                values='예산비중',
                names='매체명',
                hole=0.3,
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_pie.update_traces(
                textposition='inside',
                textinfo='label+percent',
                hovertemplate='<b>%{label}</b><br>비중: %{percent}<br>예산: %{value:.1f}%<extra></extra>'
            )
            fig_pie.update_layout(
                showlegend=True,
                height=400,
                margin=dict(t=30, b=0, l=0, r=0)
            )
            st.plotly_chart(fig_pie, use_container_width=True)
    
    with chart_col2:
        st.markdown("##### 시나리오별 전환수 비교")
        # 바 차트 - 시나리오별 전환수
        bar_data = pd.DataFrame({
            '시나리오': ['보수안(-5%)', '기본안(0%)', '공격안(+10%)'],
            '총전환수': [row['총전환수'] for row in summary_data]
        })
        
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            x=bar_data['시나리오'],
            y=bar_data['총전환수'],
            text=bar_data['총전환수'],
            texttemplate='%{text:,}건',
            textposition='outside',
            marker=dict(
                color=['#FF6B6B', '#4ECDC4', '#45B7D1'],
                line=dict(color='white', width=2)
            ),
            hovertemplate='<b>%{x}</b><br>전환수: %{y:,}건<extra></extra>'
        ))
        fig_bar.update_layout(
            yaxis_title="전환수 (건)",
            height=400,
            margin=dict(t=30, b=0, l=0, r=0),
            showlegend=False
        )
        st.plotly_chart(fig_bar, use_container_width=True)
    
    # 상세 결과
    st.markdown("---")
    st.subheader("📋 시나리오별 상세 결과")
    
    tab1, tab2, tab3 = st.tabs(["보수안(-5%)", "기본안(0%)", "공격안(+10%)"])
    
    for i, (scenario_key, tab) in enumerate(zip(['conservative', 'base', 'aggressive'], [tab1, tab2, tab3])):
        with tab:
            df = create_scenario_dataframe(scenarios[scenario_key], budget)
            
            # 데이터 포맷팅
            df_display = df.copy()
            df_display['예산'] = df_display['예산'].apply(lambda x: f"{int(x):,}" if x else "")
            df_display['예산비중'] = df_display['예산비중'].apply(lambda x: f"{x:.1f}%" if x else "")
            df_display['CPM'] = df_display['CPM'].apply(lambda x: f"{int(x):,}" if x and x != '' else "")
            df_display['예상노출'] = df_display['예상노출'].apply(lambda x: f"{int(x):,}" if x else "")
            df_display['예상클릭'] = df_display['예상클릭'].apply(lambda x: f"{int(x):,}" if x else "")
            df_display['CTR'] = df_display['CTR'].apply(lambda x: f"{x:.1f}%" if x else "")
            df_display['CPC'] = df_display['CPC'].apply(lambda x: f"{int(x):,}" if x and x != '' else "")
            df_display['CVR'] = df_display['CVR'].apply(lambda x: f"{x:.1f}%" if x else "")
            df_display['CPA'] = df_display['CPA'].apply(lambda x: f"{int(x):,}" if x else "")
            df_display['ROAS'] = df_display['ROAS'].apply(lambda x: f"{x:.1f}%" if x else "")
            
            st.dataframe(df_display, use_container_width=True, hide_index=True)
    
    # 스마트 추천
    st.markdown("---")
    st.subheader("💡 스마트 추천")
    
    recommendations = generate_recommendations(scenarios, budget)
    
    if recommendations:
        for rec in recommendations:
            if rec['type'] == 'error':
                st.error(f"{rec['icon']} {rec['message']}")
            elif rec['type'] == 'warning':
                st.warning(f"{rec['icon']} {rec['message']}")
            else:
                st.info(f"{rec['icon']} {rec['message']}")
    else:
        st.success("✅ 현재 미디어믹스가 균형잡혀 있습니다!")
    
    # AI 인사이트
    st.markdown("---")
    st.subheader("🤖 AI 인사이트")
    
    # 결과에서 업종, 월, 목표 정보 추출
    industry = results.get('industry', '보험')
    month = results.get('month', 1)
    goal = results.get('goal', '균형')
    
    # 인사이트 생성
    result_data_for_insights = {
        'scenarios': scenarios,
        'budget': budget,
        'media_list': results.get('selected_media', [])
    }
    
    insights = generate_ai_insights(result_data_for_insights, industry, month, goal)
    
    if insights:
        for insight in insights:
            if insight['type'] == 'success':
                with st.container():
                    st.success(f"**{insight['title']}**")
                    st.markdown(insight['message'])
            
            elif insight['type'] == 'error':
                with st.container():
                    st.error(f"**{insight['title']}**")
                    st.markdown(insight['message'])
            
            elif insight['type'] == 'warning':
                with st.container():
                    st.warning(f"**{insight['title']}**")
                    st.markdown(insight['message'])
            
            else:  # info
                with st.container():
                    st.info(f"**{insight['title']}**")
                    st.markdown(insight['message'])
    else:
        st.info("💡 추가 인사이트가 없습니다. 현재 설정이 양호합니다.")
    
    # 비교 모드 버튼
    st.markdown("---")
    if st.button("🔄 다른 시나리오와 비교하기", use_container_width=True):
        st.session_state.comparison_mode = True
        st.session_state.previous_result = {
            'scenarios': scenarios,
            'budget': budget,
            'industry': industry,
            'month': month,
            'goal': goal,
            'selected_media': results.get('selected_media', [])
        }
        st.info("👆 위에서 설정을 변경한 후 다시 실행하면 두 결과를 비교할 수 있습니다.")
        st.rerun()
    
    # Excel 다운로드
    st.markdown("---")
    st.subheader("📥 결과 다운로드")
    
    try:
        # 공통 함수 사용
        excel_output, filename = create_excel_download(
            scenarios=scenarios,
            budget=budget,
            mode_name="수동입력_미디어믹스",
            summary_df=summary_df
        )
        
        st.download_button(
            label="📥 Excel 결과 다운로드",
            data=excel_output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.success("✅ Excel 파일이 준비되었습니다. 다운로드 버튼을 클릭하세요.")
    
    except Exception as e:
        st.error(f"❌ Excel 생성 오류: {e}")


# 푸터 렌더링
render_footer()

